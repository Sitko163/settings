import asyncio
import logging
import random

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm  # Для обычного входа
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from flights.forms import TelegramAuthForm, TelegramCodeForm
from flights.models import User, Pilot
from telegram import Bot
from flights.utils.axes_logger import log_telegram_auth_attempt, log_telegram_code_attempt


BOT_TOKEN = settings.TOKEN

logger = logging.getLogger(__name__)


async def send_telegram_code(telegram_id, code):
    """Отправка кода в Telegram"""
    try:
        bot = Bot(token=BOT_TOKEN)
        await bot.send_message(
            chat_id=telegram_id,
            text=f"Ваш код подтверждения для входа: {code}\n\nКод действителен 5 минут."
        )
        return True
    except Exception as e:
        print(f"Ошибка отправки сообщения: {e}")
        return False


@require_http_methods(["GET", "POST"])
@csrf_protect
def login_view(request):
    """Универсальная страница входа - перенаправляет на стандартный вход"""
    if request.user.is_authenticated:
        return redirect('/')

    if request.method == 'POST':
        if 'login_type' in request.POST and request.POST['login_type'] == 'telegram':
            return telegram_login_step1_post(request)
        else:
            return standard_login_post(request)

    # Перенаправляем сразу на стандартный вход
    return redirect('standard_login')


@require_http_methods(["GET", "POST"])
@csrf_protect
def standard_login_view(request):
    """Обычный вход по username/password"""
    if request.user.is_authenticated:
        return redirect('/')
    
    if request.method == 'POST':
        return standard_login_post(request)
    
    form = AuthenticationForm()
    return render(request, 'login/login_standard.html', {'form': form})


def standard_login_post(request):
    """Обычный вход по username/password"""
    form = AuthenticationForm(request, data=request.POST)
    if form.is_valid():
        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        
        # Убрали проверку - пользователь может использовать любое имя, включая rubicon-app
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Добро пожаловать, {username}!")
            return redirect('/')
        else:
            messages.error(request, "Неверное имя пользователя или пароль. Убедитесь, что вы используете имя пользователя из Keycloak, а не название клиента.")
    else:
        # Показываем конкретные ошибки формы (но не дублируем non_field_errors, они уже показываются в шаблоне)
        if form.errors:
            for field, errors in form.errors.items():
                if field != '__all__':  # non_field_errors уже показываются в шаблоне
                    for error in errors:
                        messages.error(request, f"{field}: {error}")

    return render(request, 'login/login_standard.html', {'form': form})


def telegram_login_step1_post(request):
    """Telegram вход - шаг 1"""
    username = request.POST.get('username')
    if not username:
        messages.error(request, "Пожалуйста, введите username")
        return render(request, 'login/login_telegram.html')

    try:
        user = User.objects.get(username=username)

        # Проверяем, есть ли у пользователя связанный пилот
        if not user.pilot:
            log_telegram_auth_attempt(request, username, successful=False)
            messages.error(request, "Для этого пользователя доступен только вход по паролю")
            return render(request, 'login/login_choice.html')

        # Генерируем код
        code = str(random.randint(100000, 999999))
        request.session['auth_code'] = code
        request.session['auth_username'] = username
        request.session.set_expiry(300)  # 5 минут

        # Отправляем код в Telegram
        success = asyncio.run(send_telegram_code(user.pilot.tg_id, code))

        if success:
            messages.success(request, "Код отправлен в ваш Telegram. Введите его ниже.")
            log_telegram_auth_attempt(request, username, successful=True)
            return redirect('telegram_login_step2')
        else:
            messages.error(request, "Не удалось отправить код. Попробуйте позже.")
            log_telegram_auth_attempt(request, username, successful=False)
            return render(request, 'login/login_telegram.html')

    except User.DoesNotExist:
        log_telegram_auth_attempt(request, username, successful=False)
        messages.error(request, "Пользователь не найден")
        return render(request, 'login/login_telegram.html')
    except Exception as e:
        logger.error(f"Ошибка при входе через Telegram: {e}")
        messages.error(request, "Произошла ошибка. Попробуйте позже.")
        return render(request, 'login/login_telegram.html')


@require_http_methods(["GET", "POST"])
@csrf_protect
def telegram_login_step1(request):
    """Telegram вход - шаг 1 (GET)"""
    if request.user.is_authenticated:
        return redirect('/')

    form = TelegramAuthForm()
    return render(request, 'login/login_telegram.html', {'form': form})


@require_http_methods(["GET", "POST"])
@csrf_protect
def telegram_login_step2(request):
    """Telegram вход - шаг 2 (ввод кода)"""
    if request.user.is_authenticated:
        return redirect('/')

    if 'auth_code' not in request.session or 'auth_username' not in request.session:
        messages.error(request, "Сессия истекла. Начните заново.")
        return redirect('telegram_login_step1')

    if request.method == 'POST':
        form = TelegramCodeForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data.get('code')
            stored_code = request.session.get('auth_code')
            username = request.session.get('auth_username')

            if code == stored_code:
                try:
                    user = User.objects.get(username=username)
                    login(request, user)
                    log_telegram_code_attempt(request, username, successful=True)
                    messages.success(request, f"Добро пожаловать, {username}!")
                    # Очищаем сессию
                    del request.session['auth_code']
                    del request.session['auth_username']
                    return redirect('/')
                except User.DoesNotExist:
                    log_telegram_code_attempt(request, username, successful=False)
                    messages.error(request, "Пользователь не найден")
                    return redirect('telegram_login_step1')
            else:
                log_telegram_code_attempt(request, username, successful=False)
                messages.error(request, "Неверный код")
    else:
        form = TelegramCodeForm()

    return render(request, 'login/telegram_code.html', {'form': form})


@require_http_methods(["GET"])
def telegram_login_cancel(request):
    """Отмена входа через Telegram"""
    if 'auth_code' in request.session:
        del request.session['auth_code']
    if 'auth_username' in request.session:
        del request.session['auth_username']
    messages.info(request, "Вход отменен")
    return redirect('login')


@login_required(login_url='login')
def logout_view(request):
    """Выход из системы"""
    from django.contrib.auth import logout
    logout(request)
    messages.success(request, "Вы успешно вышли из системы")
    return redirect('login')


@login_required(login_url='login')
def debug_ip(request):
    """Отладочная функция для проверки IP"""
    ip = request.META.get('REMOTE_ADDR')
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    return JsonResponse({
        'ip': ip,
        'forwarded': forwarded,
        'user': str(request.user) if request.user.is_authenticated else 'anonymous'
    })


@login_required(login_url='login')
def map_view(request):
    context = {
        'yandex_api_key': settings.YANDEX_API_KEY,
    }
    return render(request, 'map.html', context)

@login_required(login_url='login')
def statistics_view(request):
    return render(request, 'statistics.html')

@login_required(login_url='login')
def schedule_view(request):
    return render(request, 'schedule.html')

@login_required(login_url='login')
def rating_view(request):
    return render(request, 'rating.html')

@login_required(login_url='login')
def reports_view(request):
    """Страница отчетов"""
    return render(request, 'reports.html')


@login_required(login_url='login')
def export_report_excel(request):
    """Экспорт отчета в Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.drawing.image import Image
    import matplotlib
    matplotlib.use('Agg')  # Используем backend без GUI
    import matplotlib.pyplot as plt
    import io
    import base64
    from datetime import datetime
    from flights.api.reports import ReportsDataView
    from flights.models import Flight, FlightResultTypes
    
    try:
        # Получаем данные через API
        api_view = ReportsDataView()
        api_view.request = request
        response = api_view.get(request)
        
        if response.status_code != 200:
            return HttpResponse('Ошибка получения данных', status=500)
        
        data = response.data
        
        # Создаем рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"Сводный отчет о полетах"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Общая статистика
        ws[f'A{row}'] = "Общая статистика"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        stats_headers = ['Показатель', 'Значение']
        for col, header in enumerate(stats_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        stats_data = [
            ['Всего вылетов', data['summary']['total_flights']],
            ['Уничтожено', data['summary']['destroyed_flights']],
            ['Поражено', data['summary']['defeated_flights']],
            ['Не поражено', data['summary']['not_defeated_flights']],
            ['% Уничтожения', f"{data['summary']['destruction_rate_percent']}%"],
            ['% Успеха', f"{data['summary']['success_rate_percent']}%"],
        ]
        
        for stat_row in stats_data:
            for col, value in enumerate(stat_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if col == 2:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        row += 2
        
        # Статистика по пилотам
        ws[f'A{row}'] = "Статистика по пилотам"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        pilot_headers = ['Пилот', 'Всего', 'Уничтожено', 'Поражено', 'Не поражено', '% Успеха']
        for col, header in enumerate(pilot_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        pilot_start_row = row
        for pilot in data['pilots'][:20]:  # Ограничиваем до 20 пилотов
            ws.cell(row=row, column=1).value = pilot['pilot_name']
            ws.cell(row=row, column=2).value = pilot['total_flights']
            ws.cell(row=row, column=3).value = pilot['destroyed_flights']
            ws.cell(row=row, column=4).value = pilot['defeated_flights']
            ws.cell(row=row, column=5).value = pilot['not_defeated_flights']
            ws.cell(row=row, column=6).value = f"{pilot['success_rate_percent']}%"
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Диаграмма по пилотам
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Топ-10 пилотов по количеству вылетов"
        chart.y_axis.title = 'Количество вылетов'
        chart.x_axis.title = 'Пилоты'
        
        data_ref = Reference(ws, min_col=2, min_row=pilot_start_row-1, max_row=min(row-1, pilot_start_row+9))
        cats_ref = Reference(ws, min_col=1, min_row=pilot_start_row, max_row=min(row-1, pilot_start_row+9))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 15
        
        ws.add_chart(chart, f'H{pilot_start_row}')
        
        row += 15
        
        # Статистика по целям
        ws[f'A{row}'] = "Статистика по целям"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        target_headers = ['Тип цели', 'Всего', 'Уничтожено', 'Поражено', 'Не поражено']
        for col, header in enumerate(target_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        target_start_row = row
        for target in data['targets'][:15]:  # Ограничиваем до 15 целей
            ws.cell(row=row, column=1).value = target['target']
            ws.cell(row=row, column=2).value = target['total_flights']
            ws.cell(row=row, column=3).value = target['destroyed_flights']
            ws.cell(row=row, column=4).value = target['defeated_flights']
            ws.cell(row=row, column=5).value = target['not_defeated_flights']
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Создаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в Excel: {e}", exc_info=True)
        return HttpResponse(f'Ошибка: {str(e)}', status=500)


@login_required(login_url='login')
def export_report_pdf(request):
    """Экспорт отчета в PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import io
    from datetime import datetime
    from flights.api.reports import ReportsDataView
    
    try:
        # Получаем данные через API
        api_view = ReportsDataView()
        api_view.request = request
        response = api_view.get(request)
        
        if response.status_code != 200:
            return HttpResponse('Ошибка получения данных', status=500)
        
        data = response.data
        
        # Создаем PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Заголовок
        story.append(Paragraph("Сводный отчет о полетах", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Общая статистика
        story.append(Paragraph("Общая статистика", heading_style))
        
        summary_data = [
            ['Показатель', 'Значение'],
            ['Всего вылетов', str(data['summary']['total_flights'])],
            ['Уничтожено', str(data['summary']['destroyed_flights'])],
            ['Поражено', str(data['summary']['defeated_flights'])],
            ['Не поражено', str(data['summary']['not_defeated_flights'])],
            ['% Уничтожения', f"{data['summary']['destruction_rate_percent']}%"],
            ['% Успеха', f"{data['summary']['success_rate_percent']}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Статистика по пилотам
        story.append(Paragraph("Статистика по пилотам (Топ-10)", heading_style))
        
        pilot_data = [['Пилот', 'Всего', 'Уничтожено', 'Поражено', '% Успеха']]
        for pilot in data['pilots'][:10]:
            pilot_data.append([
                pilot['pilot_name'],
                str(pilot['total_flights']),
                str(pilot['destroyed_flights']),
                str(pilot['defeated_flights']),
                f"{pilot['success_rate_percent']}%"
            ])
        
        pilot_table = Table(pilot_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        pilot_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(pilot_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Диаграмма по пилотам
        fig, ax = plt.subplots(figsize=(8, 5))
        top_pilots = data['pilots'][:10]
        pilot_names = [p['pilot_name'] for p in top_pilots]
        pilot_totals = [p['total_flights'] for p in top_pilots]
        
        ax.barh(pilot_names, pilot_totals, color='#366092')
        ax.set_xlabel('Количество вылетов')
        ax.set_title('Топ-10 пилотов по количеству вылетов')
        ax.invert_yaxis()
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        img = Image(img_buffer, width=6*inch, height=3.75*inch)
        story.append(img)
        story.append(Spacer(1, 0.3*inch))
        
        # Статистика по целям
        story.append(Paragraph("Статистика по целям (Топ-10)", heading_style))
        
        target_data = [['Тип цели', 'Всего', 'Уничтожено', 'Поражено']]
        for target in data['targets'][:10]:
            target_data.append([
                target['target'],
                str(target['total_flights']),
                str(target['destroyed_flights']),
                str(target['defeated_flights'])
            ])
        
        target_table = Table(target_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        target_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(target_table)
        
        # Строим PDF
        doc.build(story)
        
        # Возвращаем ответ
        response = HttpResponse(content_type='application/pdf')
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        return response
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в PDF: {e}", exc_info=True)
        return HttpResponse(f'Ошибка: {str(e)}', status=500)
