import datetime
import logging

from asgiref.sync import async_to_sync
from config import settings
from django import forms
from django.contrib import admin, messages
from django.contrib.admin import SimpleListFilter
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.templatetags.static import static
from django.urls import path
from django.urls import reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from flights.models import Pilot, ExplosiveDevice, ExplosiveType, Drone, TargetType, CorrectiveType, Flight, \
    FlightResultTypes, \
    FlightObjectiveTypes, User, DirectionType, DroneTypes, ImportProgress
from telegram import Bot

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


admin.AdminSite.site_header = '–ê—ç—Ä–æ—Ñ–ª–æ—Ç'
admin.AdminSite.site_title = '–ü–∞–Ω–µ–ª—å —É–ø—Ä–∞–≤–ª–µ–Ω–∏—è'
admin.AdminSite.index_title = '–ê–¥–º–∏–Ω–∏—Å—Ç—Ä–∏—Ä–æ–≤–∞–Ω–∏–µ'


def send_telegram_message(bot_token, chat_id, message):
    """–û—Ç–ø—Ä–∞–≤–∫–∞ —Å–æ–æ–±—â–µ–Ω–∏—è —á–µ—Ä–µ–∑ Telegram"""

    async def _send():
        bot = Bot(token=bot_token)
        await bot.send_message(chat_id=chat_id, text=message)

    return async_to_sync(_send)()


@admin.action(description="–û—Ç–ø—Ä–∞–≤–∏—Ç—å —Å–æ–æ–±—â–µ–Ω–∏–µ –≤ Telegram")
def send_telegram_broadcast(modeladmin, request, queryset):
    """Action –¥–ª—è –æ—Ç–ø—Ä–∞–≤–∫–∏ —Å–æ–æ–±—â–µ–Ω–∏—è –≤—ã–±—Ä–∞–Ω–Ω—ã–º –ø–∏–ª–æ—Ç–∞–º"""
    if 'apply' in request.POST:
        message = request.POST.get('message', '')
        if not message:
            modeladmin.message_user(request, "–°–æ–æ–±—â–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º!", messages.ERROR)
            return HttpResponseRedirect(request.get_full_path())

        BOT_TOKEN = settings.TOKEN

        success_count = 0
        fail_count = 0

        for pilot in queryset:
            try:
                send_telegram_message(BOT_TOKEN, pilot.tg_id, message)
                success_count += 1
            except Exception as e:
                fail_count += 1
                modeladmin.message_user(
                    request,
                    f"–û—à–∏–±–∫–∞ –æ—Ç–ø—Ä–∞–≤–∫–∏ {pilot.callname}: {str(e)}",
                    messages.WARNING
                )

        modeladmin.message_user(
            request,
            f"–°–æ–æ–±—â–µ–Ω–∏–µ –æ—Ç–ø—Ä–∞–≤–ª–µ–Ω–æ: {success_count} —É—Å–ø–µ—à–Ω–æ, {fail_count} –æ—à–∏–±–æ–∫",
            messages.SUCCESS if fail_count == 0 else messages.WARNING
        )
        return HttpResponseRedirect(request.get_full_path())

    # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º —Ñ–æ—Ä–º—É –¥–ª—è –≤–≤–æ–¥–∞ —Å–æ–æ–±—â–µ–Ω–∏—è
    context = {
        'pilots': queryset,
        'pilot_count': queryset.count(),
        'action_checkbox_name': '_selected_action',  # –ò—Å–ø—Ä–∞–≤–ª–µ–Ω–æ!
    }
    return render(request, 'admin/pilot_broadcast.html', context)


class CustomAdminSite(admin.AdminSite):
    def each_context(self, request):
        context = super().each_context(request)
        context['custom_css'] = static('admin/css/custom_admin.css')
        return context

#admin_site = CustomAdminSite(name='custom_admin')


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    fieldsets = (
        (None, {'fields': ('username', 'password')}),
        (_('Personal info'), {'fields': ('first_name', 'last_name', 'email', 'phone')}),
        (_('Permissions'), {
            'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions'),
        }),
        (_('Important dates'), {'fields': ('last_login', 'date_joined')}),
        (_('Pilot Info'), {'fields': ('pilot',)}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'password1', 'password2', 'phone', 'pilot', 'is_staff', 'is_superuser'),
        }),
    )

    # –î–æ–±–∞–≤–ª—è–µ–º –∞–≤—Ç–æ–∑–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –¥–ª—è –ø–æ–ª—è pilot
    autocomplete_fields = ['pilot']

    list_display = ('username', 'email', 'first_name', 'last_name', 'is_staff', 'get_pilot_callname')
    list_filter = ('is_staff', 'is_superuser', 'is_active', 'groups')
    search_fields = ('username', 'first_name', 'last_name', 'email', 'phone')
    ordering = ('username',)

    def get_pilot_callname(self, obj):
        if obj.pilot:
            return obj.pilot.callname
        return "-"

    get_pilot_callname.short_description = _('Pilot Callname')
    get_pilot_callname.admin_order_field = 'pilot__callname'


@admin.register(ExplosiveDevice)
class ExplosiveDeviceAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    ordering = ('name',)
    actions = ['delete_all']
    
    def delete_all(self, request, queryset):
        """–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"""
        count = ExplosiveDevice.objects.count()
        ExplosiveDevice.objects.all().delete()
        self.message_user(request, f'–£–¥–∞–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {count}', level='success')
    delete_all.short_description = "–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"

@admin.register(ExplosiveType)
class ExplosiveTypeAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    ordering = ('name',)
    actions = ['delete_all']
    
    def delete_all(self, request, queryset):
        """–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"""
        count = ExplosiveType.objects.count()
        ExplosiveType.objects.all().delete()
        self.message_user(request, f'–£–¥–∞–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {count}', level='success')
    delete_all.short_description = "–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"

@admin.register(ImportProgress)
class ImportProgressAdmin(admin.ModelAdmin):
    list_display = ('file_name', 'file_size', 'last_processed_row', 'total_rows', 'total_created', 'is_completed', 'last_import_date')
    list_filter = ('is_completed', 'last_import_date')
    search_fields = ('file_name', 'file_hash')
    readonly_fields = ('file_hash', 'last_import_date', 'created', 'modified')
    ordering = ['-last_import_date']
    
    def has_add_permission(self, request):
        return False  # –ó–∞–ø–∏—Å–∏ —Å–æ–∑–¥–∞—é—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –ø—Ä–∏ –∏–º–ø–æ—Ä—Ç–µ
    
    def has_delete_permission(self, request, obj=None):
        return True  # –ú–æ–∂–Ω–æ —É–¥–∞–ª—è—Ç—å –¥–ª—è —Å–±—Ä–æ—Å–∞ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞

@admin.register(Drone)
class DroneAdmin(admin.ModelAdmin):
    search_fields = ('name', 'drone_type')
    list_display = ('name', 'drone_type')
    ordering = ('name',)
    actions = ['delete_all']
    
    def delete_all(self, request, queryset):
        """–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"""
        count = Drone.objects.count()
        Drone.objects.all().delete()
        self.message_user(request, f'–£–¥–∞–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {count}', level='success')
    delete_all.short_description = "–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"

@admin.register(TargetType)
class TargetTypeAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    ordering = ('name',)

@admin.register(DirectionType)
class DirectionTypeAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    ordering = ('name',)

@admin.register(CorrectiveType)
class CorrectiveTypeAdmin(admin.ModelAdmin):
    search_fields = ('name',)
    ordering = ('name',)


@admin.register(Pilot)
class PilotAdmin(admin.ModelAdmin):
    list_display = (
        'callname',
        'tg_id',
        'engineer_callname',
        'driver_callname',
        'drone_type',
        'video_type',
        'manual_type',
        'flights_count',
    )
    list_filter = (
        'drone_type',
    )
    search_fields = (
        'callname',
        'tg_id', # –ü–æ–∏—Å–∫ –ø–æ Telegram ID
    )
    ordering = ('callname',)
    readonly_fields = ('id', 'created', 'modified')
    actions = [send_telegram_broadcast, 'delete_all']
    
    def delete_all(self, request, queryset):
        """–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"""
        count = Pilot.objects.count()
        Pilot.objects.all().delete()
        self.message_user(request, f'–£–¥–∞–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {count}', level='success')
    delete_all.short_description = "–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('broadcast/', self.admin_site.admin_view(self.broadcast_view), name='flights_pilot_broadcast'),
        ]
        return custom_urls + urls

    def broadcast_view(self, request):
        """–°—Ç—Ä–∞–Ω–∏—Ü–∞ –º–∞—Å—Å–æ–≤–æ–π —Ä–∞—Å—Å—ã–ª–∫–∏"""
        if request.method == 'POST':
            message = request.POST.get('message', '')
            if message:
                pilots = Pilot.objects.all()
                BOT_TOKEN = settings.TOKEN

                success_count = 0
                fail_count = 0

                for pilot in pilots:
                    try:
                        send_telegram_message(BOT_TOKEN, pilot.tg_id, message)
                        success_count += 1
                    except Exception as e:
                        fail_count += 1

                messages.success(
                    request,
                    f"–†–∞—Å—Å—ã–ª–∫–∞ –∑–∞–≤–µ—Ä—à–µ–Ω–∞: {success_count} —É—Å–ø–µ—à–Ω–æ, {fail_count} –æ—à–∏–±–æ–∫"
                )
            else:
                messages.error(request, "–°–æ–æ–±—â–µ–Ω–∏–µ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –ø—É—Å—Ç—ã–º!")

        context = {
            'pilot_count': Pilot.objects.count(),
            **self.admin_site.each_context(request),
        }
        return render(request, 'admin/broadcast_form.html', context)

    def send_message_link(self, obj):
        return format_html(
            '<a class="button" href="{}">–°–æ–æ–±—â–µ–Ω–∏–µ</a>&nbsp;'
            '<a class="button" href="{}" style="background: #28a745;">–ú–∞—Å—Å–æ–≤–∞—è —Ä–∞—Å—Å—ã–ª–∫–∞</a>',
            f"?action=send_single_message&pilot_id={obj.id}",
            reverse('admin:flights_pilot_broadcast')
        )

    send_message_link.short_description = "–î–µ–π—Å—Ç–≤–∏—è"

    def changelist_view(self, request, extra_context=None):
        if 'action' in request.GET and request.GET['action'] == 'send_single_message':
            pilot_id = request.GET.get('pilot_id')
            if pilot_id:
                # –ü–µ—Ä–µ–Ω–∞–ø—Ä–∞–≤–ª—è–µ–º –Ω–∞ action —Å —ç—Ç–∏–º –ø–∏–ª–æ—Ç–æ–º
                request.POST = request.POST.copy()
                request.POST['_selected_action'] = [pilot_id]
                request.POST['action'] = 'send_telegram_broadcast'
                request.method = 'POST'
                return self.changelist_view(request, extra_context)

        return super().changelist_view(request, extra_context)

    def flights_count(self, obj):
        count = obj.flights.count() # –ï—Å–ª–∏ related_name='flights' –≤ –º–æ–¥–µ–ª–∏ Flight
        url = reverse('admin:flights_flight_changelist') + f'?pilot__id__exact={obj.id}' # –ó–∞–º–µ–Ω–∏—Ç–µ 'flights' –Ω–∞ –∏–º—è –≤–∞—à–µ–≥–æ –ø—Ä–∏–ª–æ–∂–µ–Ω–∏—è
        return format_html('<a href="{}">{}</a>', url, count)
    flights_count.short_description = '–ö–æ–ª-–≤–æ –ø–æ–ª–µ—Ç–æ–≤'


class FlightAdminForm(forms.ModelForm):
    class Meta:
        model = Flight
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


class FlightDateFilter(SimpleListFilter):
    title = _('–î–∞—Ç–∞ –≤—ã–ª–µ—Ç–∞ (—Ç–æ—á–Ω–∞—è)')
    parameter_name = 'exact_flight_date'
    template = 'admin/date_filter.html'

    def lookups(self, request, model_admin):
        return []

    def queryset(self, request, queryset):
        date_value = request.GET.get(self.parameter_name)
        if date_value:
            try:
                from django.utils.dateparse import parse_date
                parsed_date = parse_date(date_value)
                if parsed_date:
                    return queryset.filter(flight_date=parsed_date)
            except (ValueError, TypeError):
                pass
        return queryset

    def value(self):
        return self.used_parameters.get(self.parameter_name, '')

    def has_output(self):
        return True

    def choices(self, changelist):
        yield {
            'selected': self.value() is not None,
            'query_string': changelist.get_query_string(remove=[self.parameter_name]),
            'display': _('–í—Å–µ –¥–∞—Ç—ã'),
        }

@admin.register(Flight)
class FlightAdmin(admin.ModelAdmin):
    list_display = (
        'number',
        'pilot_link',
        'drone',
        'formatted_flight_date',
        'formatted_flight_time',
        'target',
        'result_colored',
        'coordinates_preview',
        'comment_short',
        'created_display',
    )

    list_filter = (
        'flight_date',
        #FlightDateFilter,
        'pilot',
        'drone',
        'target',
        'result',
    )

    search_fields = (
        'number',
        'pilot__callname',
        'target',
        'coordinates',
    )

    fieldsets = (
        ('–û—Å–Ω–æ–≤–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è', {
            'fields': (
                'number',
                'pilot',
                'drone',
                'flight_date',
                'flight_time',
                'target',
                'comment'
            )
        }),
        ('–î–µ—Ç–∞–ª–∏ –≤—ã–ø–æ–ª–Ω–µ–Ω–∏—è', {
            'fields': (
                'engineer',
                'driver',
                'video',
                'manage',
                'distance',
                'corrective',
                'result',
                'direction',
            )
        }),
        ('–ë–æ–µ–ø—Ä–∏–ø–∞—Å—ã', {
            'fields': (
                'explosive_type',
                'explosive_device',
            ),
            'classes': ('collapse',),
        }),
        ('–¶–µ–ª—å –∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã', {
            'fields': (
                'coordinates',
                'coordinates_info_display',
                'objective',
                'drone_remains',
            )
        }),
        ('–°–ª—É–∂–µ–±–Ω–∞—è –∏–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è', {
            'fields': (
                'id',
                'created',
                'modified',
            ),
            'classes': ('collapse',),
        }),
    )

    readonly_fields = (
        'id',
        'created',
        'modified',
        'coordinates_info_display',
    )

    ordering = ('-created',)

    list_per_page = 100

    actions = ['mark_as_destroyed',
               'mark_as_defeated',
               'mark_as_not_defeated',
               'delete_all',
               'recalculate_coordinates',
               'precalculate_coordinates',
               'clear_coordinate_cache',
               'process_all_coordinates',
               ]

    change_list_template = "admin/flight_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                '<path:object_id>/recalculate-coordinates/',
                self.admin_site.admin_view(self.recalculate_coordinates_view),
                name='flights_flight_recalculate_coordinates',

            ),
            path(
                'import-xlsx/',
                self.admin_site.admin_view(self.import_xlsx_view),
                name='flights_flight_import_xlsx'),
            path(
                'clear-database/',
                self.admin_site.admin_view(self.clear_database_view),
                name='flights_flight_clear_database'),

        ]
        return custom_urls + urls

    def recalculate_coordinates_view(self, request, object_id):
        try:
            flight = self.get_object(request, object_id)
            if flight:
                flight.lat_sk42 = None
                flight.lon_sk42 = None
                flight.lat_wgs84 = None
                flight.lon_wgs84 = None
                flight.save(update_fields=[])

                coord_info = flight.get_coordinates_info_cached()

                if coord_info:
                    self.message_user(
                        request,
                        f"–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è –ø–æ–ª–µ—Ç–∞ ‚Ññ{flight.number} —É—Å–ø–µ—à–Ω–æ –ø–µ—Ä–µ—Å—á–∏—Ç–∞–Ω—ã!",
                        level=messages.SUCCESS
                    )
                else:
                    self.message_user(
                        request,
                        f"–û—à–∏–±–∫–∞ –ø–µ—Ä–µ—Å—á–µ—Ç–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –ø–æ–ª–µ—Ç–∞ ‚Ññ{flight.number}",
                        level=messages.ERROR
                    )
            else:
                self.message_user(
                    request,
                    "–ü–æ–ª–µ—Ç –Ω–µ –Ω–∞–π–¥–µ–Ω",
                    level=messages.ERROR
                )

        except Exception as e:
            self.message_user(
                request,
                f"–û—à–∏–±–∫–∞: {str(e)}",
                level=messages.ERROR
            )

        return HttpResponseRedirect(request.META.get('HTTP_REFERER', '..'))

    def mark_as_defeated(self, request, queryset):
        updated_count = queryset.update(result=FlightResultTypes.DEFEATED)
        self.message_user(request, f"{updated_count} –ø–æ–ª–µ—Ç–æ–≤ –æ—Ç–º–µ—á–µ–Ω—ã –∫–∞–∫ '–ü–æ—Ä–∞–∂–µ–Ω'.")
    mark_as_defeated.short_description = "üî• –û—Ç–º–µ—Ç–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–µ –∫–∞–∫ '–ü–æ—Ä–∞–∂–µ–Ω'"

    def mark_as_destroyed(self, request, queryset):
        updated_count = queryset.update(result=FlightResultTypes.DESTROYED)
        self.message_user(request, f"{updated_count} –ø–æ–ª–µ—Ç–æ–≤ –æ—Ç–º–µ—á–µ–Ω—ã –∫–∞–∫ '–£–Ω–∏—á—Ç–æ–∂–µ–Ω'.")
    mark_as_destroyed.short_description = "‚úÖ –û—Ç–º–µ—Ç–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–µ –∫–∞–∫ '–£–Ω–∏—á—Ç–æ–∂–µ–Ω'"

    def mark_as_not_defeated(self, request, queryset):
        updated_count = queryset.update(result=FlightResultTypes.NOT_DEFEATED)
        self.message_user(request, f"{updated_count} –ø–æ–ª–µ—Ç–æ–≤ –æ—Ç–º–µ—á–µ–Ω—ã –∫–∞–∫ '–ù–µ –ø–æ—Ä–∞–∂–µ–Ω'.")
    mark_as_not_defeated.short_description = "‚ùå –û—Ç–º–µ—Ç–∏—Ç—å –≤—ã–±—Ä–∞–Ω–Ω—ã–µ –∫–∞–∫ '–ù–µ –ø–æ—Ä–∞–∂–µ–Ω'"
    
    def delete_all(self, request, queryset):
        """–£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"""
        count = Flight.objects.count()
        Flight.objects.all().delete()
        self.message_user(request, f'–£–¥–∞–ª–µ–Ω–æ –∑–∞–ø–∏—Å–µ–π: {count}', level='success')
    delete_all.short_description = "üóëÔ∏è –£–¥–∞–ª–∏—Ç—å –≤—Å–µ –∑–∞–ø–∏—Å–∏"

    def recalculate_coordinates(self, request, queryset):
        updated_count = 0
        error_count = 0

        for flight in queryset:
            try:
                flight.lat_sk42 = None
                flight.lon_sk42 = None
                flight.lat_wgs84 = None
                flight.lon_wgs84 = None
                flight.save(update_fields=[])

                coord_info = flight.get_coordinates_info_cached()

                if coord_info:
                    updated_count += 1
                else:
                    error_count += 1

            except Exception as e:
                error_count += 1
                self.message_user(
                    request,
                    f"–û—à–∏–±–∫–∞ –ø–µ—Ä–µ—Å—á–µ—Ç–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –ø–æ–ª–µ—Ç–∞ {flight.number}: {str(e)}",
                    level=messages.ERROR
                )

        if updated_count > 0:
            self.message_user(
                request,
                f"–£—Å–ø–µ—à–Ω–æ –ø–µ—Ä–µ—Å—á–∏—Ç–∞–Ω—ã –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è {updated_count} –ø–æ–ª–µ—Ç–æ–≤.",
                level=messages.SUCCESS
            )

        if error_count > 0:
            self.message_user(
                request,
                f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–µ—Ä–µ—Å—á–µ—Ç–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è {error_count} –ø–æ–ª–µ—Ç–æ–≤.",
                level=messages.WARNING
            )

    recalculate_coordinates.short_description = "üîÑ –ü–µ—Ä–µ—Å—á–∏—Ç–∞—Ç—å –∫—ç—à–∏—Ä–æ–≤–∞–Ω–Ω—ã–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã"

    def precalculate_coordinates(self, request, queryset):
        updated_count = 0
        error_count = 0

        for flight in queryset:
            try:
                coord_info = flight.update_coordinates_from_cache()

                if coord_info:
                    updated_count += 1
                else:
                    error_count += 1

            except Exception as e:
                error_count += 1
                self.message_user(
                    request,
                    f"–û—à–∏–±–∫–∞ –ø–µ—Ä–µ—Å—á–µ—Ç–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –ø–æ–ª–µ—Ç–∞ {flight.number}: {str(e)}",
                    level=messages.ERROR
                )

        if updated_count > 0:
            self.message_user(
                request,
                f"–£—Å–ø–µ—à–Ω–æ –ø–µ—Ä–µ—Å—á–∏—Ç–∞–Ω—ã –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è {updated_count} –ø–æ–ª–µ—Ç–æ–≤.",
                level=messages.SUCCESS
            )

        if error_count > 0:
            self.message_user(
                request,
                f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø–µ—Ä–µ—Å—á–µ—Ç–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è {error_count} –ø–æ–ª–µ—Ç–æ–≤.",
                level=messages.WARNING
            )

    precalculate_coordinates.short_description = "üîÑ –ü–µ—Ä–µ—Å—á–∏—Ç–∞—Ç—å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏–∑ –∫–µ—à–∞"

    def clear_coordinate_cache(self, request, queryset):
        cleared_count = 0

        for flight in queryset:
            try:
                flight.lat_sk42 = 90.0
                flight.lon_sk42 = 0.0
                flight.lat_wgs84 = 90.0
                flight.lon_wgs84 = 0.0
                flight.save(update_fields=['lat_sk42', 'lon_sk42', 'lat_wgs84', 'lon_wgs84'])
                cleared_count += 1

            except Exception as e:
                self.message_user(
                    request,
                    f"–û—à–∏–±–∫–∞ –æ—á–∏—Å—Ç–∫–∏ –∫—ç—à–∞ –¥–ª—è –ø–æ–ª–µ—Ç–∞ {flight.number}: {str(e)}",
                    level=messages.ERROR
                )

        self.message_user(
            request,
            f"–û—á–∏—â–µ–Ω –∫—ç—à –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è {cleared_count} –ø–æ–ª–µ—Ç–æ–≤.",
            level=messages.INFO
        )

    clear_coordinate_cache.short_description = "üßπ –û—á–∏—Å—Ç–∏—Ç—å –∫—ç—à –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç"

    def recalculate_all_coordinates(self, request, queryset):
        all_flights = Flight.objects.all()
        total_count = all_flights.count()
        updated_count = 0
        error_count = 0

        self.message_user(
            request,
            f"–ù–∞—á–∏–Ω–∞–µ–º –ø–µ—Ä–µ—Å—á–µ—Ç –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –≤—Å–µ—Ö {total_count} –ø–æ–ª–µ—Ç–æ–≤...",
            level=messages.INFO
        )

        for flight in all_flights:
            try:
                flight.lat_sk42 = None
                flight.lon_sk42 = None
                flight.lat_wgs84 = None
                flight.lon_wgs84 = None
                flight.save(update_fields=[])

                coord_info = flight.get_coordinates_info_cached()

                if coord_info:
                    updated_count += 1
                else:
                    error_count += 1

                if (updated_count + error_count) % 100 == 0:
                    self.message_user(
                        request,
                        f"–û–±—Ä–∞–±–æ—Ç–∞–Ω–æ {updated_count + error_count} –∏–∑ {total_count} –ø–æ–ª–µ—Ç–æ–≤...",
                        level=messages.INFO
                    )

            except Exception as e:
                error_count += 1
                continue

        self.message_user(
            request,
            f"–ü–µ—Ä–µ—Å—á–µ—Ç –∑–∞–≤–µ—Ä—à–µ–Ω! –£—Å–ø–µ—à–Ω–æ: {updated_count}, –û—à–∏–±–æ–∫: {error_count}",
            level=messages.SUCCESS if error_count == 0 else messages.WARNING
        )

    recalculate_all_coordinates.short_description = "üîÑ –ü–µ—Ä–µ—Å—á–∏—Ç–∞—Ç—å –í–°–ï –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã"

    def process_all_coordinates(self, request, queryset):
        """–û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –≤—Å–µ—Ö –ø–æ–ª–µ—Ç–æ–≤ —Å –Ω–µ–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–º–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏"""
        # –ü–æ–ª—É—á–∞–µ–º –≤—Å–µ –ø–æ–ª–µ—Ç—ã —Å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏, –∫–æ—Ç–æ—Ä—ã–µ –µ—â–µ –Ω–µ –æ–±—Ä–∞–±–æ—Ç–∞–Ω—ã
        flights_to_process = Flight.objects.filter(
            coordinates__isnull=False
        ).exclude(
            coordinates=''
        ).filter(
            lat_wgs84__isnull=True
        )
        
        total_count = flights_to_process.count()
        if total_count == 0:
            self.message_user(
                request,
                "–ù–µ—Ç –ø–æ–ª–µ—Ç–æ–≤ —Å –Ω–µ–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–º–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏.",
                level=messages.INFO
            )
            return
        
        self.message_user(
            request,
            f"–ù–∞–π–¥–µ–Ω–æ {total_count} –ø–æ–ª–µ—Ç–æ–≤ —Å –Ω–µ–æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã–º–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏. –ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É (—ç—Ç–æ –º–æ–∂–µ—Ç –∑–∞–Ω—è—Ç—å –Ω–µ–∫–æ—Ç–æ—Ä–æ–µ –≤—Ä–µ–º—è)...",
            level=messages.INFO
        )
        
        def progress_callback(processed, total):
            """Callback –¥–ª—è –æ—Ç—Å–ª–µ–∂–∏–≤–∞–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞"""
            if processed % 1000 == 0 or processed == total:
                self.message_user(
                    request,
                    f"–û–±—Ä–∞–±–æ—Ç–∞–Ω–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç: {processed}/{total}",
                    level=messages.INFO
                )
        
        try:
            success_count, error_count = Flight.batch_process_coordinates(
                queryset=flights_to_process,
                batch_size=500,
                update_callback=progress_callback
            )
            
            if success_count > 0:
                self.message_user(
                    request,
                    f"‚úì –£—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è {success_count} –ø–æ–ª–µ—Ç–æ–≤.",
                    level=messages.SUCCESS
                )
            
            if error_count > 0:
                self.message_user(
                    request,
                    f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è {error_count} –ø–æ–ª–µ—Ç–æ–≤.",
                    level=messages.WARNING
                )
        except Exception as e:
            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç: {e}", exc_info=True)
            self.message_user(
                request,
                f"‚ùå –ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –æ–±—Ä–∞–±–æ—Ç–∫–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç: {str(e)}",
                level=messages.ERROR
            )
            return
        
        # –û—á–∏—â–∞–µ–º –∫—ç—à API –∫–∞—Ä—Ç—ã
        from django.core.cache import cache
        try:
            if hasattr(cache, 'delete_pattern'):
                cache.delete_pattern('rubicon:flights_total:*')
            else:
                cache.clear()
            logger.info("–ö—ç—à –¥–ª—è API –∫–∞—Ä—Ç—ã –æ—á–∏—â–µ–Ω –ø–æ—Å–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç")
        except Exception as cache_error:
            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –æ—á–∏—Å—Ç–∏—Ç—å –∫—ç—à: {cache_error}")

    process_all_coordinates.short_description = "üó∫Ô∏è –û–±—Ä–∞–±–æ—Ç–∞—Ç—å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –¥–ª—è –≤—Å–µ—Ö –ø–æ–ª–µ—Ç–æ–≤ –±–µ–∑ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç"

    def pilot_link(self, obj):
        if obj.pilot:
            url = reverse('admin:flights_pilot_change', args=[obj.pilot.id])
            return format_html('<a href="{}">{}</a>', url, obj.pilot.callname)
        return "-"
    pilot_link.short_description = '–ü–∏–ª–æ—Ç'
    pilot_link.admin_order_field = 'pilot__callname'

    def formatted_flight_date(self, obj):
        if obj.flight_date:
            return obj.flight_date.strftime('%d.%m.%Y')
        return "-"
    formatted_flight_date.short_description = '–î–∞—Ç–∞'
    formatted_flight_date.admin_order_field = 'flight_date'

    def formatted_flight_time(self, obj):
        if obj.flight_time:
            return obj.flight_time.strftime('%H:%M')
        return "-"
    formatted_flight_time.short_description = '–í—Ä–µ–º—è'
    formatted_flight_time.admin_order_field = 'flight_time'

    def created_display(self, obj):
        if obj.created:
            return obj.created.strftime('%d.%m.%Y %H:%M')
        return "-"

    created_display.short_description = '–°–æ–∑–¥–∞–Ω'
    created_display.admin_order_field = 'created'

    def modified_display(self, obj):
        if obj.modified:
            return obj.modified.strftime('%d.%m.%Y %H:%M')
        return "-"

    modified_display.short_description = '–ò–∑–º–µ–Ω–µ–Ω'
    modified_display.admin_order_field = 'modified'

    def result_colored(self, obj):
        if obj.result == FlightResultTypes.DESTROYED:
            color = 'green'
            text = '‚úÖ –£–Ω–∏—á—Ç–æ–∂–µ–Ω–æ'
        elif obj.result == FlightResultTypes.NOT_DEFEATED:
            color = 'red'
            text = '‚ùå –ù–µ –ø–æ—Ä–∞–∂–µ–Ω–æ'
        elif obj.result == FlightResultTypes.DEFEATED:
            color = 'orange'
            text = 'üî• –ü–æ—Ä–∞–∂–µ–Ω–æ'
        else:
            color = 'gray'
            text = obj.result
        return format_html('<span style="color: {};">{}</span>', color, text)
    result_colored.short_description = '–†–µ–∑—É–ª—å—Ç–∞—Ç'
    result_colored.admin_order_field = 'result'

    def objective_colored(self, obj):
        if obj.objective == FlightObjectiveTypes.EXISTS:
            color = 'blue'
            text = '–ï—Å—Ç—å'
        elif obj.objective == FlightObjectiveTypes.NOT_EXISTS:
            color = 'gray'
            text = '–ù–µ—Ç'
        else:
            color = 'black'
            text = obj.objective
        return format_html('<span style="color: {};">{}</span>', color, text)
    objective_colored.short_description = '–û–±—ä–µ–∫—Ç–∏–≤'
    objective_colored.admin_order_field = 'objective'

    def coordinates_preview(self, obj):
        if obj.coordinates:
            preview = (obj.coordinates[:20] + '...') if len(obj.coordinates) > 20 else obj.coordinates
            coord_info = obj.get_coordinates_info_cached()
            if coord_info and coord_info.get('lat_wgs84') and coord_info.get('lon_wgs84'):
                lat = coord_info['lat_wgs84']
                lon = coord_info['lon_wgs84']
                map_url = f"https://www.google.com/maps?q={lat},{lon}"
                return format_html(
                    '<span title="{}">{}</span>',
                    f"–°–ö-42: {obj.coordinates}",
                    preview
                )
            else:
                return format_html('<span title="{}">{}</span>', f"–°–ö-42: {obj.coordinates}", preview)
        return "-"
    coordinates_preview.short_description = '–ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã (–°–ö-42)'

    def comment_short(self, obj):
        if obj.comment:
            return (obj.comment[:20] + '...') if len(obj.comment) > 20 else obj.comment
        return "-"
    comment_short.short_description = '–ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π'

    def coordinates_info_display(self, obj):
        coord_info = obj.get_coordinates_info_cached() # –ò—Å–ø–æ–ª—å–∑—É–µ–º –∫—ç—à–∏—Ä–æ–≤–∞–Ω–Ω—É—é –≤–µ—Ä—Å–∏—é
        if coord_info:
            try:
                lat_sk42 = coord_info.get('lat_sk42', 'N/A')
                lon_sk42 = coord_info.get('lon_sk42', 'N/A')
                lat_wgs84 = coord_info.get('lat_wgs84', 'N/A')
                lon_wgs84 = coord_info.get('lon_wgs84', 'N/A')

                return format_html(
                    "<strong>–°–ö-42 (–≥—Ä–∞–¥—É—Å—ã):</strong> —à–∏—Ä–æ—Ç–∞: {}, –¥–æ–ª–≥–æ—Ç–∞: {}<br>"
                    "<strong>WGS-84 (–≥—Ä–∞–¥—É—Å—ã):</strong> —à–∏—Ä–æ—Ç–∞: {}, –¥–æ–ª–≥–æ—Ç–∞: {}",
                    round(lat_sk42, 6) if lat_sk42 != 'N/A' else 'N/A',
                    round(lon_sk42, 6) if lon_sk42 != 'N/A' else 'N/A',
                    round(lat_wgs84, 6) if lat_wgs84 != 'N/A' else 'N/A',
                    round(lon_wgs84, 6) if lon_wgs84 != 'N/A' else 'N/A'
                )
            except Exception as e:
                return f"–û—à–∏–±–∫–∞ –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è: {e}"
        else:
            try:
                temp_info = obj.get_coordinates_info() # –ù–µ –∫—ç—à–∏—Ä–æ–≤–∞–Ω–Ω–∞—è –≤–µ—Ä—Å–∏—è
                if temp_info:
                    lat_sk42 = temp_info.get('lat_sk42', 'N/A')
                    lon_sk42 = temp_info.get('lon_sk42', 'N/A')
                    lat_wgs84 = temp_info.get('lat_wgs84', 'N/A')
                    lon_wgs84 = temp_info.get('lon_wgs84', 'N/A')
                    return format_html(
                        "<strong>–°–ö-42 (–≥—Ä–∞–¥—É—Å—ã):</strong> —à–∏—Ä–æ—Ç–∞: {}, –¥–æ–ª–≥–æ—Ç–∞: {}<br>"
                        "<strong>WGS-84 (–≥—Ä–∞–¥—É—Å—ã):</strong> —à–∏—Ä–æ—Ç–∞: {}, –¥–æ–ª–≥–æ—Ç–∞: {}",
                        round(lat_sk42, 6) if lat_sk42 != 'N/A' else 'N/A',
                        round(lon_sk42, 6) if lon_sk42 != 'N/A' else 'N/A',
                        round(lat_wgs84, 6) if lat_wgs84 != 'N/A' else 'N/A',
                        round(lon_wgs84, 6) if lon_wgs84 != 'N/A' else 'N/A'
                    )
                else:
                     return "–ò–Ω—Ñ–æ—Ä–º–∞—Ü–∏—è –Ω–µ–¥–æ—Å—Ç—É–ø–Ω–∞"
            except Exception as e:
                return f"–û—à–∏–±–∫–∞ –ø–æ–ª—É—á–µ–Ω–∏—è –¥–∞–Ω–Ω—ã—Ö: {e}"
    coordinates_info_display.short_description = '–ü—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–Ω—ã–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã'

    def import_xlsx_view(self, request):
        xlsx_files = request.FILES.getlist("xlsx_files")

        if not xlsx_files:
            self.message_user(request, _("–ü–æ–∂–∞–ª—É–π—Å—Ç–∞, –≤—ã–±–µ—Ä–∏—Ç–µ —Ö–æ—Ç—è –±—ã –æ–¥–∏–Ω .xlsx –∏–ª–∏ .xlsm —Ñ–∞–π–ª."), level=messages.ERROR)
            return HttpResponseRedirect("../")

        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –Ω–µ–∑–∞–≤–µ—Ä—à–µ–Ω–Ω—ã–µ –∏–º–ø–æ—Ä—Ç—ã
        # –ï—Å–ª–∏ –µ—Å—Ç—å - –Ω–µ –æ—á–∏—â–∞–µ–º –±–∞–∑—É, –ø—Ä–æ–¥–æ–ª–∂–∞–µ–º –∏–º–ø–æ—Ä—Ç
        # –ï—Å–ª–∏ –Ω–µ—Ç - –æ—á–∏—â–∞–µ–º –±–∞–∑—É –ø–µ—Ä–µ–¥ –∏–º–ø–æ—Ä—Ç–æ–º
        has_unfinished_imports = ImportProgress.objects.filter(is_completed=False).exists()
        
        if not has_unfinished_imports:
            # –û—á–∏—â–∞–µ–º –±–∞–∑—É –≤—ã–ª–µ—Ç–æ–≤ –ø–µ—Ä–µ–¥ –∏–º–ø–æ—Ä—Ç–æ–º (—Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∏ –Ω–µ —Ç—Ä–æ–≥–∞–µ–º)
            self.message_user(request, _("–û—á–∏—Å—Ç–∫–∞ –±–∞–∑—ã –≤—ã–ª–µ—Ç–æ–≤ –ø–µ—Ä–µ–¥ –∏–º–ø–æ—Ä—Ç–æ–º..."), level=messages.INFO)
            flights_count_before = Flight.objects.count()
            Flight.objects.all().delete()
            self.message_user(request, 
                              _(f"–£–¥–∞–ª–µ–Ω–æ {flights_count_before} –∑–∞–ø–∏—Å–µ–π –≤—ã–ª–µ—Ç–æ–≤. –ù–∞—á–∏–Ω–∞–µ–º –∏–º–ø–æ—Ä—Ç..."), 
                              level=messages.SUCCESS)
        else:
            self.message_user(request, 
                              _("–ù–∞–π–¥–µ–Ω—ã –Ω–µ–∑–∞–≤–µ—Ä—à–µ–Ω–Ω—ã–µ –∏–º–ø–æ—Ä—Ç—ã. –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –∏–º–ø–æ—Ä—Ç —Å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã—Ö –ø–æ–∑–∏—Ü–∏–π..."), 
                              level=messages.INFO)

        for xlsx_file in xlsx_files:
            if not (xlsx_file.name.endswith('.xlsx') or xlsx_file.name.endswith('.xlsm')):
                self.message_user(request, _(f"–§–∞–π–ª '{xlsx_file.name}' –Ω–µ —è–≤–ª—è–µ—Ç—Å—è .xlsx –∏–ª–∏ .xlsm —Ñ–∞–π–ª–æ–º."),
                                  level=messages.ERROR)
                return HttpResponseRedirect("../")

            total_created = 0
            total_errors = []

            for xlsx_file in xlsx_files:
                try:
                    logger.info(f"–ù–∞—á–∏–Ω–∞–µ–º –∏–º–ø–æ—Ä—Ç –∏–∑ —Ñ–∞–π–ª–∞: {xlsx_file.name}, —Ä–∞–∑–º–µ—Ä: {xlsx_file.size} –±–∞–π—Ç")
                    self.message_user(request, _(f"–ù–∞—á–∏–Ω–∞–µ–º –∏–º–ø–æ—Ä—Ç –∏–∑ —Ñ–∞–π–ª–∞: {xlsx_file.name}"), level=messages.INFO)

                    # –í—ã—á–∏—Å–ª—è–µ–º hash —Ñ–∞–π–ª–∞ –¥–ª—è –∏–¥–µ–Ω—Ç–∏—Ñ–∏–∫–∞—Ü–∏–∏
                    import hashlib
                    xlsx_file.seek(0)  # –°–±—Ä–∞—Å—ã–≤–∞–µ–º –ø–æ–∑–∏—Ü–∏—é —Ñ–∞–π–ª–∞
                    file_content = xlsx_file.read()
                    file_hash = hashlib.md5(file_content).hexdigest()
                    xlsx_file.seek(0)  # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –ø–æ–∑–∏—Ü–∏—é –¥–ª—è —á—Ç–µ–Ω–∏—è
                    
                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã–π –ø—Ä–æ–≥—Ä–µ—Å—Å –¥–ª—è —ç—Ç–æ–≥–æ —Ñ–∞–π–ª–∞
                    import_progress = None
                    try:
                        import_progress = ImportProgress.objects.filter(
                            file_name=xlsx_file.name,
                            file_hash=file_hash
                        ).first()
                    except Exception as e:
                        logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∏–º–ø–æ—Ä—Ç–∞: {e}")
                    
                    start_row = 5  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é –Ω–∞—á–∏–Ω–∞–µ–º —Å —Å—Ç—Ä–æ–∫–∏ 5
                    if import_progress and not import_progress.is_completed:
                        start_row = import_progress.last_processed_row + 1
                        self.message_user(request,
                                          _(f"üìå –ù–∞–π–¥–µ–Ω —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã–π –ø—Ä–æ–≥—Ä–µ—Å—Å! –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –∏–º–ø–æ—Ä—Ç —Å —Å—Ç—Ä–æ–∫–∏ {start_row} (–±—ã–ª–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ {import_progress.last_processed_row} –∏–∑ {import_progress.total_rows})"),
                                          level=messages.INFO)
                        logger.info(f"–ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –∏–º–ø–æ—Ä—Ç —Ñ–∞–π–ª–∞ '{xlsx_file.name}' —Å —Å—Ç—Ä–æ–∫–∏ {start_row} (hash: {file_hash})")
                    elif import_progress and import_progress.is_completed:
                        self.message_user(request,
                                          _(f"‚úì –§–∞–π–ª '{xlsx_file.name}' —É–∂–µ –ø–æ–ª–Ω–æ—Å—Ç—å—é –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω. –ü—Ä–æ–ø—É—Å–∫–∞–µ–º."),
                                          level=messages.INFO)
                        logger.info(f"–§–∞–π–ª '{xlsx_file.name}' —É–∂–µ –ø–æ–ª–Ω–æ—Å—Ç—å—é –∏–º–ø–æ—Ä—Ç–∏—Ä–æ–≤–∞–Ω, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º")
                        continue
                    else:
                        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é –∑–∞–ø–∏—Å—å –ø—Ä–æ–≥—Ä–µ—Å—Å–∞
                        import_progress = ImportProgress.objects.create(
                            file_name=xlsx_file.name,
                            file_size=xlsx_file.size,
                            file_hash=file_hash,
                            last_processed_row=start_row - 1,
                            total_rows=0,  # –ë—É–¥–µ—Ç –æ–±–Ω–æ–≤–ª–µ–Ω–æ –ø–æ—Å–ª–µ –æ–ø—Ä–µ–¥–µ–ª–µ–Ω–∏—è
                            total_created=0,
                            is_completed=False
                        )
                        logger.info(f"–°–æ–∑–¥–∞–Ω–∞ –Ω–æ–≤–∞—è –∑–∞–ø–∏—Å—å –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –¥–ª—è —Ñ–∞–π–ª–∞ '{xlsx_file.name}' (hash: {file_hash})")

                    import openpyxl
                    # –ù–ï –∏—Å–ø–æ–ª—å–∑—É–µ–º read_only —Ä–µ–∂–∏–º, —Ç–∞–∫ –∫–∞–∫ –æ–Ω –Ω–µ –ø–æ–∑–≤–æ–ª—è–µ—Ç —á–∏—Ç–∞—Ç—å —è—á–µ–π–∫–∏ –≤ –ø—Ä–æ–∏–∑–≤–æ–ª—å–Ω–æ–º –ø–æ—Ä—è–¥–∫–µ
                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –æ–±—ã—á–Ω—ã–π —Ä–µ–∂–∏–º –¥–ª—è –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç–∏ —á—Ç–µ–Ω–∏—è —è—á–µ–µ–∫ —á–µ—Ä–µ–∑ ws.cell()
                    # –≠—Ç–æ –ø–æ—Ç—Ä–µ–±—É–µ—Ç –±–æ–ª—å—à–µ –ø–∞–º—è—Ç–∏, –Ω–æ –æ–±–µ—Å–ø–µ—á–∏—Ç —Å—Ç–∞–±–∏–ª—å–Ω—É—é —Ä–∞–±–æ—Ç—É
                    try:
                        if xlsx_file.name.endswith('.xlsm'):
                            wb = openpyxl.load_workbook(xlsx_file, data_only=True, keep_vba=True)
                        else:
                            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                        logger.info(f"–§–∞–π–ª '{xlsx_file.name}' –æ—Ç–∫—Ä—ã—Ç –≤ –æ–±—ã—á–Ω–æ–º —Ä–µ–∂–∏–º–µ (–¥–ª—è —Å—Ç–∞–±–∏–ª—å–Ω–æ–π —Ä–∞–±–æ—Ç—ã —Å –±–æ–ª—å—à–∏–º–∏ —Ñ–∞–π–ª–∞–º–∏)")
                    except Exception as e:
                        # –ï—Å–ª–∏ keep_vba –Ω–µ –ø–æ–¥–¥–µ—Ä–∂–∏–≤–∞–µ—Ç—Å—è, –ø—Ä–æ–±—É–µ–º –±–µ–∑ –Ω–µ–≥–æ
                        logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—Ç–∫—Ä—ã—Ç–∏–∏ —Ñ–∞–π–ª–∞ '{xlsx_file.name}' —Å keep_vba: {e}, –ø—Ä–æ–±—É–µ–º –±–µ–∑ keep_vba")
                        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
                    
                    # –ò—â–µ–º –ª–∏—Å—Ç "–°–í–û–î–ù–ê–Ø"
                    sheet_name = None
                    for sheet in wb.sheetnames:
                        if sheet.upper() == 'SVODNAYA' or '–°–í–û–î–ù–ê–Ø' in sheet.upper():
                            sheet_name = sheet
                            break
                    
                    if sheet_name is None:
                        self.message_user(request,
                                          _(f"–õ–∏—Å—Ç '–°–í–û–î–ù–ê–Ø' –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Ñ–∞–π–ª–µ '{xlsx_file.name}'."),
                                          level=messages.ERROR)
                        continue
                    
                    ws = wb[sheet_name]

                    # –°—Ç—Ä—É–∫—Ç—É—Ä–∞ —Ñ–∞–π–ª–∞: –∑–∞–≥–æ–ª–æ–≤–∫–∏ –≤ —Å—Ç—Ä–æ–∫–µ 4 (B4-V4), –¥–∞–Ω–Ω—ã–µ —Å —Å—Ç—Ä–æ–∫–∏ 5
                    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–æ–ª–æ–Ω–∫–∏ –ø–æ –æ–ø–∏—Å–∞–Ω–∏—é (openpyxl –∏—Å–ø–æ–ª—å–∑—É–µ—Ç 1-based –∏–Ω–¥–µ–∫—Å—ã):
                    # B=2, C=3, D=4, E=5, F=6, H=8, I=9, J=10, K=11, L=12, M=13, O=15, Q=17, R=18, S=19, T=20, U=21, V=22
                    COL_TIME = 2  # B - –≤—Ä–µ–º—è
                    COL_TARGET = 3  # C - —Ö–∞—Ä–∞–∫—Ç–µ—Ä —Ü–µ–ª–∏
                    COL_COMMENT = 4  # D - –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π
                    COL_COORD_X = 5  # E - –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞ X
                    COL_COORD_Y = 6  # F - –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞ Y
                    COL_DRONE = 8  # H - —Ç–∏–ø –¥—Ä–æ–Ω–∞
                    COL_DRONE_COUNT = 9  # I - –∫–æ–ª-–≤–æ –¥—Ä–æ–Ω–æ–≤ (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –º–æ–¥–µ–ª–∏)
                    COL_EXPLOSIVE_TYPE = 10  # J - –≤–∏–¥ –ë–ü
                    COL_EXPLOSIVE_USAGE = 11  # K - —Ä–∞—Å—Ö–æ–¥ –ë–ü (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –º–æ–¥–µ–ª–∏)
                    COL_EXPLOSIVE_DEVICE = 12  # L - –≤–∏–¥ –≤–∑—Ä—ã–≤–∞—Ç–µ–ª—è
                    COL_EXPLOSIVE_DEVICE_COUNT = 13  # M - –∫–æ–ª-–≤–æ –≤–∑—Ä—ã–≤–∞—Ç–µ–ª—è (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –º–æ–¥–µ–ª–∏)
                    COL_RESULT = 15  # O - —Ä–µ–∑—É–ª—å—Ç–∞—Ç –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è
                    COL_CALCULATION_NUMBER = 17  # Q - –Ω–æ–º–µ—Ä —Ä–∞—Å—á–µ—Ç–∞ (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –º–æ–¥–µ–ª–∏)
                    COL_FLIGHT_DATE = 18  # R - –¥–∞—Ç–∞ –≤—ã–ª–µ—Ç–∞
                    COL_FLIGHT_NUMBER = 19  # S - –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞
                    COL_DISTANCE = 20  # T - –¥–∏—Å—Ç–∞–Ω—Ü–∏—è
                    COL_FLIGHT_TIME_DURATION = 21  # U - –≤—Ä–µ–º—è –ø–æ–ª–µ—Ç–∞ (–Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è –≤ –º–æ–¥–µ–ª–∏, –µ—Å—Ç—å flight_time)
                    COL_OPERATOR_CALLNAME = 22  # V - –ø–æ–∑—ã–≤–Ω–æ–π –æ–ø–µ—Ä–∞—Ç–æ—Ä–∞

                    created_count = 0
                    error_messages = []
                    
                    # –ü—Ä–µ–¥–∑–∞–≥—Ä—É–∂–∞–µ–º —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∏ –≤ –ø–∞–º—è—Ç—å –¥–ª—è –±—ã—Å—Ç—Ä–æ–≥–æ –¥–æ—Å—Ç—É–ø–∞
                    self.message_user(request, _("–ü—Ä–µ–¥–∑–∞–≥—Ä—É–∑–∫–∞ —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–æ–≤..."), level=messages.INFO)
                    
                    # –ö—ç—à –ø–∏–ª–æ—Ç–æ–≤ –ø–æ –ø–æ–∑—ã–≤–Ω–æ–º—É
                    pilots_cache = {pilot.callname.lower(): pilot for pilot in Pilot.objects.all()}
                    
                    # –ö—ç—à —Ç–∏–ø–æ–≤ —Ü–µ–ª–µ–π
                    target_types_cache = {tt.name.lower(): tt for tt in TargetType.objects.all()}
                    
                    # –ö—ç—à —Ç–∏–ø–æ–≤ –¥—Ä–æ–Ω–æ–≤
                    # –°–æ–∑–¥–∞–µ–º –∫—ç—à –¥—Ä–æ–Ω–æ–≤ —Å –∫–ª—é—á–æ–º –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤ –¥–ª—è –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∏ X-51 –∏ X51
                    def get_drone_comparison_key(drone_name):
                        """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –¥—Ä–æ–Ω–æ–≤ (–±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤)"""
                        return re.sub(r'[-]', '', str(drone_name).lower().strip())
                    drones_cache = {get_drone_comparison_key(drone.name): drone for drone in Drone.objects.all()}
                    
                    # –ö—ç—à –≤–∏–¥–æ–≤ –ë–ü
                    explosive_types_cache = {et.name.lower(): et for et in ExplosiveType.objects.all()}
                    
                    # –ö—ç—à –≤–∏–¥–æ–≤ –≤–∑—Ä—ã–≤–∞—Ç–µ–ª–µ–π
                    explosive_devices_cache = {ed.name.lower(): ed for ed in ExplosiveDevice.objects.all()}
                    
                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–µ –ø–æ–ª–µ—Ç—ã –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –¥—É–±–ª–∏–∫–∞—Ç–æ–≤
                    # –°–æ–∑–¥–∞–µ–º –º–Ω–æ–∂–µ—Å—Ç–≤–æ –∫–ª—é—á–µ–π —Å—É—â–µ—Å—Ç–≤—É—é—â–∏—Ö –ø–æ–ª–µ—Ç–æ–≤: (number, pilot_id, flight_date, flight_time)
                    existing_flights_set = set(
                        Flight.objects.values_list('number', 'pilot_id', 'flight_date', 'flight_time')
                    )
                    logger.info(f"–ù–∞–π–¥–µ–Ω–æ {len(existing_flights_set)} —Å—É—â–µ—Å—Ç–≤—É—é—â–∏—Ö –ø–æ–ª–µ—Ç–æ–≤ –≤ –ë–î")
                    
                    # –°–ø–∏—Å–∫–∏ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è –Ω–æ–≤—ã—Ö –∑–∞–ø–∏—Å–µ–π –≤ —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∞—Ö
                    new_target_types = {}
                    new_drones = {}
                    new_pilots = {}
                    new_explosive_types = {}
                    new_explosive_devices = {}
                    
                    # –ò–Ω–∏—Ü–∏–∞–ª–∏–∑–∏—Ä—É–µ–º —Å–ø–∏—Å–∫–∏ –¥–ª—è bulk –æ–ø–µ—Ä–∞—Ü–∏–π
                    flights_to_create = []
                    # –ö—ç—à –¥—É–±–ª–∏–∫–∞—Ç–æ–≤ –±–æ–ª—å—à–µ –Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è - –±–∞–∑–∞ –æ—á–∏—â–µ–Ω–∞, –≤—Å–µ –∑–∞–ø–∏—Å–∏ –Ω–æ–≤—ã–µ
                    
                    # –§—É–Ω–∫—Ü–∏–∏ –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏–∏ –¥–ª—è —Ü–µ–ª–µ–π –∏ –¥—Ä–æ–Ω–æ–≤ (—á—Ç–æ–±—ã –∏–∑–±–µ–∂–∞—Ç—å –¥—É–±–ª–∏—Ä–æ–≤–∞–Ω–∏—è)
                    import re
                    
                    def normalize_target_name(target_name):
                        """–ù–æ—Ä–º–∞–ª–∏–∑—É–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ —Ü–µ–ª–∏ –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è –¥—É–±–ª–∏–∫–∞—Ç–æ–≤"""
                        if not target_name:
                            return None
                        target_str = str(target_name).strip()
                        if not target_str:
                            return None
                        
                        target_lower = target_str.lower()
                        # –£–¥–∞–ª—è–µ–º –ª–∏—à–Ω–∏–µ –ø—Ä–æ–±–µ–ª—ã, –¥–µ—Ñ–∏—Å—ã, —Ç–æ—á–∫–∏, –∑–∞–ø—è—Ç—ã–µ
                        target_normalized = re.sub(r'[-\s\.\,]+', ' ', target_lower).strip()
                        
                        # –û–±—ä–µ–¥–∏–Ω—è–µ–º –ø–æ—Ö–æ–∂–∏–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã
                        # –ê–≤—Ç–æ–º–æ–±–∏–ª—å–Ω–∞—è —Ç–µ—Ö–Ω–∏–∫–∞
                        if any(word in target_normalized for word in ['–∞–≤—Ç–æ–º–æ–±–∏–ª—å–Ω', '–∞–≤—Ç–æ—Ç–µ—Ö–Ω–∏–∫', '–∞–≤—Ç–æ —Ç–µ—Ö–Ω–∏–∫', '–∞–≤—Ç–æ-—Ç–µ—Ö–Ω–∏–∫']):
                            return '–ê–≤—Ç–æ–º–æ–±–∏–ª—å–Ω–∞—è —Ç–µ—Ö–Ω–∏–∫–∞'
                        # –ü–í–•
                        if '–ø–≤—Ö' in target_normalized:
                            match = re.search(r'–ø–≤—Ö\s*[-\s]*(\d+[–∏]?)', target_normalized, re.IGNORECASE)
                            if match:
                                return f"–ü–í–•-{match.group(1).upper()}"
                            return '–ü–í–•'
                        
                        # –û–±—â–∞—è –Ω–æ—Ä–º–∞–ª–∏–∑–∞—Ü–∏—è
                        normalized = re.sub(r'[^\w\s]', '', target_str)
                        normalized = ' '.join(normalized.split())
                        if normalized:
                            normalized = normalized[0].upper() + normalized[1:].lower() if len(normalized) > 1 else normalized.upper()
                        return normalized if normalized else None
                    
                    def normalize_drone_name(drone_name):
                        """–ù–æ—Ä–º–∞–ª–∏–∑—É–µ—Ç –Ω–∞–∑–≤–∞–Ω–∏–µ –¥—Ä–æ–Ω–∞ –¥–ª—è –æ–±—ä–µ–¥–∏–Ω–µ–Ω–∏—è –¥—É–±–ª–∏–∫–∞—Ç–æ–≤.
                        –°–æ—Ö—Ä–∞–Ω—è–µ—Ç –¥–µ—Ñ–∏—Å—ã –≤ –Ω–∞–∑–≤–∞–Ω–∏–∏, –Ω–æ –≥—Ä—É–ø–ø–∏—Ä—É–µ—Ç X-51 –∏ X51 –∫–∞–∫ –æ–¥–∏–Ω –¥—Ä–æ–Ω."""
                        if not drone_name:
                            return None
                        drone_str = str(drone_name).strip()
                        if not drone_str:
                            return None
                        
                        drone_lower = drone_str.lower()
                        # –£–¥–∞–ª—è–µ–º –ª–∏—à–Ω–∏–µ –ø—Ä–æ–±–µ–ª—ã, —Ç–æ—á–∫–∏, –∑–∞–ø—è—Ç—ã–µ (–Ω–æ –ù–ï –¥–µ—Ñ–∏—Å—ã!)
                        drone_normalized = re.sub(r'[\s\.\,]+', ' ', drone_lower).strip()
                        
                        # –û–±—ä–µ–¥–∏–Ω—è–µ–º –ø–æ—Ö–æ–∂–∏–µ –≤–∞—Ä–∏–∞–Ω—Ç—ã
                        # –ü–í–•
                        if '–ø–≤—Ö' in drone_normalized:
                            match = re.search(r'–ø–≤—Ö\s*[-]?\s*(\d+[–∏]?)', drone_normalized, re.IGNORECASE)
                            if match:
                                # –í—Å–µ–≥–¥–∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –≤–µ—Ä—Å–∏—é —Å –¥–µ—Ñ–∏—Å–æ–º
                                return f"–ü–í–•-{match.group(1).upper()}"
                            return '–ü–í–•'
                        # –ú–æ–ª–Ω–∏—è
                        if '–º–æ–ª–Ω–∏—è' in drone_normalized:
                            match = re.search(r'–º–æ–ª–Ω–∏—è\s*[-]?\s*(\d+[–¥—Ç]?)', drone_normalized, re.IGNORECASE)
                            if match:
                                # –í—Å–µ–≥–¥–∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –≤–µ—Ä—Å–∏—é —Å –¥–µ—Ñ–∏—Å–æ–º
                                return f"–ú–æ–ª–Ω–∏—è-{match.group(1).upper()}"
                            return '–ú–æ–ª–Ω–∏—è'
                        # –ö–í–ù - –∏—â–µ–º –ö–í–ù —Å –Ω–æ–º–µ—Ä–æ–º –∏–ª–∏ –±—É–∫–≤–æ–π (–ö–í–ù-–¢, –ö–í–ù-23, –ö–í–ù-23–¢ –∏ —Ç.–¥.)
                        if '–∫–≤–Ω' in drone_normalized:
                            # –ü–∞—Ç—Ç–µ—Ä–Ω: –ö–í–ù + –æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω—ã–π –¥–µ—Ñ–∏—Å + —Ü–∏—Ñ—Ä—ã + –æ–ø—Ü–∏–æ–Ω–∞–ª—å–Ω–∞—è –±—É–∫–≤–∞ (–¢, —Ç –∏ —Ç.–¥.)
                            match = re.search(r'–∫–≤–Ω\s*[-]?\s*(\d+[—Ç—Ç]?|[—Ç—Ç]|\d+)', drone_normalized, re.IGNORECASE)
                            if match:
                                suffix = match.group(1).upper()
                                # –í—Å–µ–≥–¥–∞ –≤–æ–∑–≤—Ä–∞—â–∞–µ–º –≤–µ—Ä—Å–∏—é —Å –¥–µ—Ñ–∏—Å–æ–º
                                return f"–ö–í–ù-{suffix}"
                            return '–ö–í–ù'
                        
                        # –î–ª—è –æ—Å—Ç–∞–ª—å–Ω—ã—Ö - –Ω–æ—Ä–º–∞–ª–∏–∑—É–µ–º –ø—Ä–æ–±–µ–ª—ã, –Ω–æ —Å–æ—Ö—Ä–∞–Ω—è–µ–º –¥–µ—Ñ–∏—Å—ã
                        # –ü—Ä–∏–≤–æ–¥–∏–º –∫ –ø—Ä–∞–≤–∏–ª—å–Ω–æ–º—É —Ä–µ–≥–∏—Å—Ç—Ä—É: –ø–µ—Ä–≤–∞—è –±—É–∫–≤–∞ –∑–∞–≥–ª–∞–≤–Ω–∞—è, –æ—Å—Ç–∞–ª—å–Ω—ã–µ —Å—Ç—Ä–æ—á–Ω—ã–µ
                        normalized = re.sub(r'\s+', ' ', drone_str.strip())
                        if normalized:
                            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –¥–µ—Ñ–∏—Å—ã –∏ –±—É–∫–≤–µ–Ω–Ω–æ-—Ü–∏—Ñ—Ä–æ–≤—ã–µ —Å–∏–º–≤–æ–ª—ã
                            # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ —Å–ª–æ–≤–∞ –ø–æ –ø—Ä–æ–±–µ–ª–∞–º –∏ –¥–µ—Ñ–∏—Å–∞–º, –Ω–æ —Å–æ—Ö—Ä–∞–Ω—è–µ–º —Å—Ç—Ä—É–∫—Ç—É—Ä—É
                            words = re.split(r'([-\s])', normalized)
                            result = ''
                            for word in words:
                                if word in ['-', ' ']:
                                    result += word
                                elif word:
                                    # –ü–µ—Ä–≤–∞—è –±—É–∫–≤–∞ –∑–∞–≥–ª–∞–≤–Ω–∞—è, –æ—Å—Ç–∞–ª—å–Ω—ã–µ —Å—Ç—Ä–æ—á–Ω—ã–µ
                                    result += word[0].upper() + word[1:].lower() if len(word) > 1 else word.upper()
                            return result if result else None
                        return None
                    
                    # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∫–∞–∂–¥—ã–µ 1000 —Å—Ç—Ä–æ–∫
                    progress_interval = 1000

                    # –î–∞–Ω–Ω—ã–µ –Ω–∞—á–∏–Ω–∞—é—Ç—Å—è —Å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω–æ–π —Å—Ç—Ä–æ–∫–∏ –∏–ª–∏ —Å —Å—Ç—Ä–æ–∫–∏ 5 (–∑–∞–≥–æ–ª–æ–≤–∫–∏ –≤ —Å—Ç—Ä–æ–∫–µ 4)
                    data_start_row = start_row
                    # –ù–µ –æ–≥—Ä–∞–Ω–∏—á–∏–≤–∞–µ–º data_end_row - –±—É–¥–µ–º –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å –¥–æ –∫–æ–Ω—Ü–∞ —Ñ–∞–π–ª–∞ –∏–ª–∏ –¥–æ 30 –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥
                    data_end_row = ws.max_row
                    
                    # –û–±–Ω–æ–≤–ª—è–µ–º total_rows –≤ import_progress
                    if import_progress:
                        import_progress.total_rows = data_end_row
                        import_progress.save(update_fields=['total_rows'])
                    
                    # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –∫–æ–Ω—Å—Ç–∞–Ω—Ç—É –¥–ª—è –º–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–≥–æ –∫–æ–ª–∏—á–µ—Å—Ç–≤–∞ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥
                    # –£–≤–µ–ª–∏—á–µ–Ω–æ –¥–æ 50000, —á—Ç–æ–±—ã –Ω–µ –æ—Å—Ç–∞–Ω–∞–≤–ª–∏–≤–∞—Ç—å—Å—è –Ω–∞ –±–æ–ª—å—à–∏—Ö —Ä–∞–∑—Ä—ã–≤–∞—Ö –¥–∞–Ω–Ω—ã—Ö
                    # –ï—Å–ª–∏ –ø–æ—Å–ª–µ 50000 –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥ –≤—Å–µ –µ—â–µ –Ω–µ—Ç –¥–∞–Ω–Ω—ã—Ö, –∑–Ω–∞—á–∏—Ç —Ñ–∞–π–ª –∑–∞–∫–æ–Ω—á–∏–ª—Å—è
                    MAX_EMPTY_ROWS = 50000  # –£–≤–µ–ª–∏—á–µ–Ω–æ –¥–ª—è –æ—á–µ–Ω—å –±–æ–ª—å—à–∏—Ö —Ñ–∞–π–ª–æ–≤ —Å —Ä–∞–∑—Ä—ã–≤–∞–º–∏
                    CHECK_AHEAD_ROWS = 50000  # –°–∫–æ–ª—å–∫–æ —Å—Ç—Ä–æ–∫ –ø—Ä–æ–≤–µ—Ä—è—Ç—å –ø–æ—Å–ª–µ MAX_EMPTY_ROWS
                    
                    self.message_user(request, 
                                      _(f"–ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É —Å —Å—Ç—Ä–æ–∫–∏ {data_start_row} –¥–æ –∫–æ–Ω—Ü–∞ —Ñ–∞–π–ª–∞ (–∏–ª–∏ –¥–æ {MAX_EMPTY_ROWS} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥). –í—Å–µ–≥–æ —Å—Ç—Ä–æ–∫ –≤ —Ñ–∞–π–ª–µ: {ws.max_row}"),
                                      level=messages.INFO)

                    # –ü–æ–ª—É—á–∞–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ —è—á–µ–π–∫–∏ –Ω–∞–ø—Ä—è–º—É—é (–æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏—è)
                    # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –≤—Å–µ —Å—Ç—Ä–æ–∫–∏ (bulk –æ–ø–µ—Ä–∞—Ü–∏–∏ –ø–æ–∑–≤–æ–ª—è—é—Ç –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å –±–æ–ª—å—à–∏–µ –æ–±—ä–µ–º—ã)
                    total_rows_to_process = data_end_row - data_start_row + 1
                    self.message_user(request,
                                      _(f"–ù–∞—á–∏–Ω–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É {total_rows_to_process} —Å—Ç—Ä–æ–∫..."),
                                      level=messages.INFO)
                    
                    processed_row_count = 0
                    empty_rows_count = 0  # –°—á–µ—Ç—á–∏–∫ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥
                    skipped_no_flight_number = 0  # –°—á–µ—Ç—á–∏–∫ –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ –±–µ–∑ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞
                    skipped_errors = 0  # –°—á–µ—Ç—á–∏–∫ –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ –∏–∑-–∑–∞ –æ—à–∏–±–æ–∫
                    skipped_no_date = 0  # –°—á–µ—Ç—á–∏–∫ –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ –±–µ–∑ –¥–∞—Ç—ã
                    skipped_no_pilot = 0  # –°—á–µ—Ç—á–∏–∫ –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ –±–µ–∑ –ø–∏–ª–æ—Ç–∞
                    processed_successfully = 0  # –°—á–µ—Ç—á–∏–∫ —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫
                    skipped_duplicates = 0  # –°—á–µ—Ç—á–∏–∫ –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö –¥—É–±–ª–∏–∫–∞—Ç–æ–≤
                    
                    for row_idx in range(data_start_row, data_end_row + 1):
                        processed_row_count += 1
                        
                        # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∏ —Å–æ—Ö—Ä–∞–Ω—è–µ–º –µ–≥–æ –∫–∞–∂–¥—ã–µ 1000 –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫
                        if processed_row_count % 1000 == 0:
                            progress_percent = processed_row_count * 100 // total_rows_to_process
                            self.message_user(request,
                                              _(f"üìä –û–±—Ä–∞–±–æ—Ç–∞–Ω–æ —Å—Ç—Ä–æ–∫: {processed_row_count} –∏–∑ {total_rows_to_process} ({progress_percent}%). –°–æ–∑–¥–∞–Ω–æ –∑–∞–ø–∏—Å–µ–π: {created_count}, —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ: {processed_successfully}"),
                                              level=messages.INFO)
                            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∫–∞–∂–¥—ã–µ 1000 —Å—Ç—Ä–æ–∫ –¥–ª—è –≤–æ–∑–º–æ–∂–Ω–æ—Å—Ç–∏ –ø—Ä–æ–¥–æ–ª–∂–µ–Ω–∏—è –ø—Ä–∏ —Å–±–æ–µ
                            if import_progress:
                                try:
                                    import_progress.last_processed_row = data_start_row + processed_row_count - 1
                                    import_progress.total_created = created_count
                                    import_progress.save(update_fields=['last_processed_row', 'total_created', 'last_import_date'])
                                    logger.debug(f"–ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω: —Å—Ç—Ä–æ–∫–∞ {import_progress.last_processed_row}, —Å–æ–∑–¥–∞–Ω–æ {created_count}")
                                except Exception as progress_error:
                                    logger.warning(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∏–º–ø–æ—Ä—Ç–∞: {progress_error}")
                        # –ò—Å–ø–æ–ª—å–∑—É–µ–º iter_rows –¥–ª—è –±–æ–ª–µ–µ –±—ã—Å—Ç—Ä–æ–≥–æ —á—Ç–µ–Ω–∏—è (—Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏)
                        # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –Ω–∞–ø—Ä—è–º—É—é
                        row_values = [None] * (COL_OPERATOR_CALLNAME + 1)
                        
                        # –ß–∏—Ç–∞–µ–º —Ç–æ–ª—å–∫–æ –Ω—É–∂–Ω—ã–µ –∫–æ–ª–æ–Ω–∫–∏ –Ω–∞–ø—Ä—è–º—É—é (–æ–ø—Ç–∏–º–∏–∑–∏—Ä–æ–≤–∞–Ω–æ)
                        try:
                            # –ò—Å–ø–æ–ª—å–∑—É–µ–º –ø—Ä—è–º–æ–µ —á—Ç–µ–Ω–∏–µ —è—á–µ–µ–∫ —á–µ—Ä–µ–∑ cell() –¥–ª—è –Ω—É–∂–Ω—ã—Ö –∫–æ–ª–æ–Ω–æ–∫
                            # –≠—Ç–æ –±—ã—Å—Ç—Ä–µ–µ —á–µ–º —á–∏—Ç–∞—Ç—å –≤—Å—é —Å—Ç—Ä–æ–∫—É
                            row_values[COL_TIME - 1] = ws.cell(row=row_idx, column=COL_TIME).value
                            row_values[COL_TARGET - 1] = ws.cell(row=row_idx, column=COL_TARGET).value
                            row_values[COL_COMMENT - 1] = ws.cell(row=row_idx, column=COL_COMMENT).value
                            row_values[COL_COORD_X - 1] = ws.cell(row=row_idx, column=COL_COORD_X).value
                            row_values[COL_COORD_Y - 1] = ws.cell(row=row_idx, column=COL_COORD_Y).value
                            row_values[COL_DRONE - 1] = ws.cell(row=row_idx, column=COL_DRONE).value
                            row_values[COL_EXPLOSIVE_TYPE - 1] = ws.cell(row=row_idx, column=COL_EXPLOSIVE_TYPE).value
                            row_values[COL_EXPLOSIVE_DEVICE - 1] = ws.cell(row=row_idx, column=COL_EXPLOSIVE_DEVICE).value
                            row_values[COL_RESULT - 1] = ws.cell(row=row_idx, column=COL_RESULT).value
                            row_values[COL_FLIGHT_DATE - 1] = ws.cell(row=row_idx, column=COL_FLIGHT_DATE).value
                            row_values[COL_FLIGHT_NUMBER - 1] = ws.cell(row=row_idx, column=COL_FLIGHT_NUMBER).value
                            row_values[COL_DISTANCE - 1] = ws.cell(row=row_idx, column=COL_DISTANCE).value
                            row_values[COL_OPERATOR_CALLNAME - 1] = ws.cell(row=row_idx, column=COL_OPERATOR_CALLNAME).value
                        except Exception as row_error:
                            # –ï—Å–ª–∏ –Ω–µ —É–¥–∞–ª–æ—Å—å –ø—Ä–æ—á–∏—Ç–∞—Ç—å —Å—Ç—Ä–æ–∫—É, –ø—Ä–æ–ø—É—Å–∫–∞–µ–º
                            logger.warning(f"–û—à–∏–±–∫–∞ —á—Ç–µ–Ω–∏—è —Å—Ç—Ä–æ–∫–∏ {row_idx}: {row_error}")
                            continue

                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –¥–∞–Ω–Ω—ã–µ –≤ —Å—Ç—Ä–æ–∫–µ
                        # –°—Ç—Ä–æ–∫–∞ —Å—á–∏—Ç–∞–µ—Ç—Å—è –ø—É—Å—Ç–æ–π —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –Ω–µ—Ç –¥–∞—Ç—ã, –ø–æ–∑—ã–≤–Ω–æ–≥–æ –ò —Ç–∏–ø–∞ –¥—Ä–æ–Ω–∞ (–º–∏–Ω–∏–º–∞–ª—å–Ω—ã–µ —Ç—Ä–µ–±–æ–≤–∞–Ω–∏—è)
                        time_value = row_values[COL_TIME - 1] if len(row_values) >= COL_TIME else None
                        pilot_value = row_values[COL_OPERATOR_CALLNAME - 1] if len(row_values) >= COL_OPERATOR_CALLNAME else None
                        date_value = row_values[COL_FLIGHT_DATE - 1] if len(row_values) >= COL_FLIGHT_DATE else None
                        drone_value = row_values[COL_DRONE - 1] if len(row_values) >= COL_DRONE else None
                        
                        # –ù–æ—Ä–º–∞–ª–∏–∑—É–µ–º –∑–Ω–∞—á–µ–Ω–∏—è (—É–±–∏—Ä–∞–µ–º –ø—Ä–æ–±–µ–ª—ã, –ø—Ä–æ–≤–µ—Ä—è–µ–º –Ω–∞ None –∏ –ø—É—Å—Ç–æ—Ç—É)
                        def is_empty_value(val):
                            if val is None:
                                return True
                            val_str = str(val).strip()
                            return not val_str or val_str == "" or val_str.lower() == "none" or val_str.lower() == "null"
                        
                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —è–≤–ª—è–µ—Ç—Å—è –ª–∏ —Å—Ç—Ä–æ–∫–∞ –ø—É—Å—Ç–æ–π
                        # –°—Ç—Ä–æ–∫–∞ —Å—á–∏—Ç–∞–µ—Ç—Å—è –ø—É—Å—Ç–æ–π —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –Ω–µ—Ç –í–°–ï–• –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–æ–ª–µ–π –æ–¥–Ω–æ–≤—Ä–µ–º–µ–Ω–Ω–æ
                        # –ù–æ –ø—Ä–æ–≤–µ—Ä—è–µ–º –±–æ–ª–µ–µ –º—è–≥–∫–æ - –µ—Å–ª–∏ –µ—Å—Ç—å —Ö–æ—Ç—è –±—ã –æ–¥–Ω–æ –∏–∑ –ø–æ–ª–µ–π (–¥–∞—Ç–∞, –ø–∏–ª–æ—Ç, –¥—Ä–æ–Ω, –≤—Ä–µ–º—è, –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã), —Ç–æ —Å—Ç—Ä–æ–∫–∞ –Ω–µ –ø—É—Å—Ç–∞—è
                        flight_number_value = row_values[COL_FLIGHT_NUMBER - 1] if len(row_values) >= COL_FLIGHT_NUMBER else None
                        coord_x_value = row_values[COL_COORD_X - 1] if len(row_values) >= COL_COORD_X else None
                        coord_y_value = row_values[COL_COORD_Y - 1] if len(row_values) >= COL_COORD_Y else None
                        target_value = row_values[COL_TARGET - 1] if len(row_values) >= COL_TARGET else None
                        
                        # –°—Ç—Ä–æ–∫–∞ –ù–ï –ø—É—Å—Ç–∞—è, –µ—Å–ª–∏ –µ—Å—Ç—å —Ö–æ—Ç—è –±—ã –æ–¥–Ω–æ –∏–∑ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏—Ö –ø–æ–ª–µ–π (–≤–∫–ª—é—á–∞—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏ —Ü–µ–ª—å)
                        has_data = not is_empty_value(date_value) or \
                                  not is_empty_value(pilot_value) or \
                                  not is_empty_value(drone_value) or \
                                  not is_empty_value(time_value) or \
                                  not is_empty_value(flight_number_value) or \
                                  not is_empty_value(coord_x_value) or \
                                  not is_empty_value(coord_y_value) or \
                                  not is_empty_value(target_value)
                        
                        is_empty = not has_data
                        
                        # –õ–æ–≥–∏—Ä—É–µ–º –∫–∞–∂–¥—É—é 5000-—é —Å—Ç—Ä–æ–∫—É –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏ (—á—Ç–æ–±—ã –≤–∏–¥–µ—Ç—å –ø—Ä–æ–≥—Ä–µ—Å—Å)
                        if processed_row_count % 5000 == 0:
                            logger.info(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –ø—É—Å—Ç–∞—è={is_empty}, –¥–∞—Ç–∞={date_value}, –ø–∏–ª–æ—Ç={pilot_value}, –¥—Ä–æ–Ω={drone_value}, –≤—Ä–µ–º—è={time_value}, –Ω–æ–º–µ—Ä={flight_number_value}, —Å–æ–∑–¥–∞–Ω–æ={created_count}")

                        if is_empty:
                            # –ü—É—Å—Ç–∞—è —Å—Ç—Ä–æ–∫–∞ - —É–≤–µ–ª–∏—á–∏–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫
                            empty_rows_count += 1
                            
                            # –õ–æ–≥–∏—Ä—É–µ–º –∫–∞–∂–¥—ã–µ 1000 –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏ (—á—Ç–æ–±—ã –Ω–µ —Å–ø–∞–º–∏—Ç—å)
                            if empty_rows_count % 1000 == 0:
                                logger.info(f"–û–±–Ω–∞—Ä—É–∂–µ–Ω–æ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥ (—Å—Ç—Ä–æ–∫–∞ {row_idx})")
                                self.message_user(request,
                                                  _(f"‚ö†Ô∏è –û–±–Ω–∞—Ä—É–∂–µ–Ω–æ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {row_idx}..."),
                                                  level=messages.WARNING)
                            
                            # –ï—Å–ª–∏ –ø–æ–¥—Ä—è–¥ MAX_EMPTY_ROWS –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ - –ø—Ä–æ–≤–µ—Ä—è–µ–º, –µ—Å—Ç—å –ª–∏ –¥–∞–ª—å—à–µ –¥–∞–Ω–Ω—ã–µ
                            if empty_rows_count >= MAX_EMPTY_ROWS:
                                # –£–ª—É—á—à–µ–Ω–Ω–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞: –ø—Ä–æ–≤–µ—Ä—è–µ–º –±–æ–ª–µ–µ —Ç—â–∞—Ç–µ–ª—å–Ω–æ –∏ –¥–∞–ª—å—à–µ
                                has_data_anywhere = False
                                max_check = min(row_idx + CHECK_AHEAD_ROWS, data_end_row + 1)
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—É—é —Å—Ç—Ä–æ–∫—É –≤ –ø–µ—Ä–≤—ã—Ö 5000 —Å—Ç—Ä–æ–∫–∞—Ö (—É–≤–µ–ª–∏—á–µ–Ω–æ –¥–ª—è –±–æ–ª—å—à–∏—Ö —Ñ–∞–π–ª–æ–≤)
                                for check_row in range(row_idx + 1, min(row_idx + 5001, max_check)):
                                    try:
                                        check_flight_num = ws.cell(row=check_row, column=COL_FLIGHT_NUMBER).value
                                        check_time = ws.cell(row=check_row, column=COL_TIME).value
                                        check_pilot = ws.cell(row=check_row, column=COL_OPERATOR_CALLNAME).value
                                        check_date = ws.cell(row=check_row, column=COL_FLIGHT_DATE).value
                                        check_drone = ws.cell(row=check_row, column=COL_DRONE).value
                                        check_target = ws.cell(row=check_row, column=COL_TARGET).value
                                        check_coord_x = ws.cell(row=check_row, column=COL_COORD_X).value
                                        check_coord_y = ws.cell(row=check_row, column=COL_COORD_Y).value
                                        
                                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤—Å–µ –∫–ª—é—á–µ–≤—ã–µ –ø–æ–ª—è
                                        if (check_flight_num is not None and str(check_flight_num).strip() and str(check_flight_num).strip().lower() not in ['none', 'null']) or \
                                           (check_time is not None and str(check_time).strip() and str(check_time).strip().lower() not in ['none', 'null']) or \
                                           (check_pilot is not None and str(check_pilot).strip() and str(check_pilot).strip().lower() not in ['none', 'null']) or \
                                           (check_date is not None and str(check_date).strip() and str(check_date).strip().lower() not in ['none', 'null']) or \
                                           (check_drone is not None and str(check_drone).strip() and str(check_drone).strip().lower() not in ['none', 'null']) or \
                                           (check_target is not None and str(check_target).strip() and str(check_target).strip().lower() not in ['none', 'null']) or \
                                           (check_coord_x is not None and str(check_coord_x).strip() and str(check_coord_x).strip().lower() not in ['none', 'null']) or \
                                           (check_coord_y is not None and str(check_coord_y).strip() and str(check_coord_y).strip().lower() not in ['none', 'null']):
                                            has_data_anywhere = True
                                            logger.info(f"–ù–∞–π–¥–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {check_row} –ø–æ—Å–ª–µ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ (—Ç–µ–∫—É—â–∞—è —Å—Ç—Ä–æ–∫–∞: {row_idx})")
                                            break
                                    except Exception as e:
                                        logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ —Å—Ç—Ä–æ–∫–∏ {check_row}: {e}")
                                        continue
                                
                                # –ï—Å–ª–∏ –≤ –ø–µ—Ä–≤—ã—Ö 5000 —Å—Ç—Ä–æ–∫–∞—Ö –¥–∞–Ω–Ω—ã—Ö –Ω–µ—Ç, –ø—Ä–æ–≤–µ—Ä—è–µ–º –¥–∞–ª—å—à–µ –±–æ–ª–µ–µ —Ç—â–∞—Ç–µ–ª—å–Ω–æ:
                                # - —Å–ª–µ–¥—É—é—â–∏–µ 10000 —Å—Ç—Ä–æ–∫ –ø—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—ã–µ 50 —Å—Ç—Ä–æ–∫
                                # - —Å–ª–µ–¥—É—é—â–∏–µ 100000 —Å—Ç—Ä–æ–∫ –ø—Ä–æ–≤–µ—Ä—è–µ–º –∫–∞–∂–¥—ã–µ 100 —Å—Ç—Ä–æ–∫  
                                # - –¥–∞–ª—å—à–µ –∫–∞–∂–¥—ã–µ 500 —Å—Ç—Ä–æ–∫ –¥–æ –∫–æ–Ω—Ü–∞ —Ñ–∞–π–ª–∞
                                if not has_data_anywhere:
                                    # –°–ª–µ–¥—É—é—â–∏–µ 10000 —Å—Ç—Ä–æ–∫ - –∫–∞–∂–¥—ã–µ 50
                                    next_range_end = min(row_idx + 15001, max_check)
                                    for check_row in range(row_idx + 5001, next_range_end, 50):
                                        try:
                                            check_flight_num = ws.cell(row=check_row, column=COL_FLIGHT_NUMBER).value
                                            check_time = ws.cell(row=check_row, column=COL_TIME).value
                                            check_pilot = ws.cell(row=check_row, column=COL_OPERATOR_CALLNAME).value
                                            check_date = ws.cell(row=check_row, column=COL_FLIGHT_DATE).value
                                            check_drone = ws.cell(row=check_row, column=COL_DRONE).value
                                            
                                            if (check_flight_num is not None and str(check_flight_num).strip() and str(check_flight_num).strip().lower() not in ['none', 'null']) or \
                                               (check_time is not None and str(check_time).strip() and str(check_time).strip().lower() not in ['none', 'null']) or \
                                               (check_pilot is not None and str(check_pilot).strip() and str(check_pilot).strip().lower() not in ['none', 'null']) or \
                                               (check_date is not None and str(check_date).strip() and str(check_date).strip().lower() not in ['none', 'null']) or \
                                               (check_drone is not None and str(check_drone).strip() and str(check_drone).strip().lower() not in ['none', 'null']):
                                                has_data_anywhere = True
                                                logger.info(f"–ù–∞–π–¥–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {check_row} –ø–æ—Å–ª–µ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ (—Ç–µ–∫—É—â–∞—è —Å—Ç—Ä–æ–∫–∞: {row_idx})")
                                                break
                                        except Exception as e:
                                            logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ —Å—Ç—Ä–æ–∫–∏ {check_row}: {e}")
                                            continue
                                    
                                    # –ï—Å–ª–∏ –≤—Å–µ –µ—â–µ –Ω–µ –Ω–∞—à–ª–∏, –ø—Ä–æ–≤–µ—Ä—è–µ–º –¥–∞–ª—å—à–µ –∫–∞–∂–¥—ã–µ 100 —Å—Ç—Ä–æ–∫
                                    if not has_data_anywhere:
                                        next_range_end = min(row_idx + 110001, max_check)
                                        for check_row in range(row_idx + 15001, next_range_end, 100):
                                            try:
                                                check_flight_num = ws.cell(row=check_row, column=COL_FLIGHT_NUMBER).value
                                                check_time = ws.cell(row=check_row, column=COL_TIME).value
                                                check_pilot = ws.cell(row=check_row, column=COL_OPERATOR_CALLNAME).value
                                                check_date = ws.cell(row=check_row, column=COL_FLIGHT_DATE).value
                                                check_drone = ws.cell(row=check_row, column=COL_DRONE).value
                                                
                                                if (check_flight_num is not None and str(check_flight_num).strip() and str(check_flight_num).strip().lower() not in ['none', 'null']) or \
                                                   (check_time is not None and str(check_time).strip() and str(check_time).strip().lower() not in ['none', 'null']) or \
                                                   (check_pilot is not None and str(check_pilot).strip() and str(check_pilot).strip().lower() not in ['none', 'null']) or \
                                                   (check_date is not None and str(check_date).strip() and str(check_date).strip().lower() not in ['none', 'null']) or \
                                                   (check_drone is not None and str(check_drone).strip() and str(check_drone).strip().lower() not in ['none', 'null']):
                                                    has_data_anywhere = True
                                                    logger.info(f"–ù–∞–π–¥–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {check_row} –ø–æ—Å–ª–µ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ (—Ç–µ–∫—É—â–∞—è —Å—Ç—Ä–æ–∫–∞: {row_idx})")
                                                    break
                                            except Exception as e:
                                                logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ —Å—Ç—Ä–æ–∫–∏ {check_row}: {e}")
                                                continue
                                    
                                    # –ï—Å–ª–∏ –≤—Å–µ –µ—â–µ –Ω–µ –Ω–∞—à–ª–∏, –ø—Ä–æ–≤–µ—Ä—è–µ–º –¥–∞–ª—å—à–µ –∫–∞–∂–¥—ã–µ 500 —Å—Ç—Ä–æ–∫ –¥–æ –∫–æ–Ω—Ü–∞ —Ñ–∞–π–ª–∞
                                    if not has_data_anywhere:
                                        for check_row in range(row_idx + 110001, max_check, 500):
                                            try:
                                                check_flight_num = ws.cell(row=check_row, column=COL_FLIGHT_NUMBER).value
                                                check_time = ws.cell(row=check_row, column=COL_TIME).value
                                                check_pilot = ws.cell(row=check_row, column=COL_OPERATOR_CALLNAME).value
                                                check_date = ws.cell(row=check_row, column=COL_FLIGHT_DATE).value
                                                check_drone = ws.cell(row=check_row, column=COL_DRONE).value
                                                
                                                if (check_flight_num is not None and str(check_flight_num).strip() and str(check_flight_num).strip().lower() not in ['none', 'null']) or \
                                                   (check_time is not None and str(check_time).strip() and str(check_time).strip().lower() not in ['none', 'null']) or \
                                                   (check_pilot is not None and str(check_pilot).strip() and str(check_pilot).strip().lower() not in ['none', 'null']) or \
                                                   (check_date is not None and str(check_date).strip() and str(check_date).strip().lower() not in ['none', 'null']) or \
                                                   (check_drone is not None and str(check_drone).strip() and str(check_drone).strip().lower() not in ['none', 'null']):
                                                    has_data_anywhere = True
                                                    logger.info(f"–ù–∞–π–¥–µ–Ω—ã –¥–∞–Ω–Ω—ã–µ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {check_row} –ø–æ—Å–ª–µ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ (—Ç–µ–∫—É—â–∞—è —Å—Ç—Ä–æ–∫–∞: {row_idx})")
                                                    break
                                            except Exception as e:
                                                logger.debug(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ —Å—Ç—Ä–æ–∫–∏ {check_row}: {e}")
                                                continue
                                
                                if not has_data_anywhere:
                                    self.message_user(request,
                                                      _(f"‚ö†Ô∏è –û–±–Ω–∞—Ä—É–∂–µ–Ω–æ {MAX_EMPTY_ROWS}+ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥. –ü—Ä–æ–≤–µ—Ä–µ–Ω–æ –µ—â–µ –¥–æ {max_check} —Å—Ç—Ä–æ–∫–∏ - –¥–∞–Ω–Ω—ã—Ö –Ω–µ—Ç. –û–±—Ä–∞–±–æ—Ç–∫–∞ —Ñ–∞–π–ª–∞ –ø—Ä–µ–∫—Ä–∞—â–µ–Ω–∞ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {row_idx}."),
                                                      level=messages.WARNING)
                                    logger.info(f"–û—Å—Ç–∞–Ω–æ–≤–∫–∞ –∏–º–ø–æ—Ä—Ç–∞: {MAX_EMPTY_ROWS}+ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {row_idx}, –¥–∞–Ω–Ω—ã—Ö –¥–∞–ª—å—à–µ –Ω–µ—Ç (–ø—Ä–æ–≤–µ—Ä–µ–Ω–æ –¥–æ —Å—Ç—Ä–æ–∫–∏ {max_check})")
                                    break
                                else:
                                    # –ï—Å—Ç—å –¥–∞–Ω–Ω—ã–µ –¥–∞–ª—å—à–µ - —Å–±—Ä–∞—Å—ã–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫ –∏ –ø—Ä–æ–¥–æ–ª–∂–∞–µ–º
                                    self.message_user(request,
                                                      _(f"‚ö†Ô∏è –û–±–Ω–∞—Ä—É–∂–µ–Ω–æ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥, –Ω–æ –¥–∞–ª—å—à–µ –µ—Å—Ç—å –¥–∞–Ω–Ω—ã–µ (–Ω–∞–π–¥–µ–Ω–æ –Ω–∞ —Å—Ç—Ä–æ–∫–µ {check_row}). –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É..."),
                                                      level=messages.INFO)
                                    empty_rows_count = 0
                            continue
                        
                        # –ï—Å–ª–∏ —Å—Ç—Ä–æ–∫–∞ –Ω–µ –ø—É—Å—Ç–∞—è - —Å–±—Ä–∞—Å—ã–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫
                        if empty_rows_count > 0:
                            logger.debug(f"–°–±—Ä–æ—Å —Å—á–µ—Ç—á–∏–∫–∞ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ (–±—ã–ª–æ {empty_rows_count}) –Ω–∞ —Å—Ç—Ä–æ–∫–µ {row_idx}")
                            empty_rows_count = 0
                        else:
                            # –°–±—Ä–∞—Å—ã–≤–∞–µ–º —Å—á–µ—Ç—á–∏–∫, –µ—Å–ª–∏ —Å—Ç—Ä–æ–∫–∞ –Ω–µ –ø—É—Å—Ç–∞—è
                            empty_rows_count = 0

                        try:
                            # –ù–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ S (–Ω–µ–æ–±—è–∑–∞—Ç–µ–ª—å–Ω–æ–µ –ø–æ–ª–µ)
                            flight_number_raw = row_values[COL_FLIGHT_NUMBER - 1] if len(row_values) >= COL_FLIGHT_NUMBER else None
                            flight_number = None
                            
                            if flight_number_raw is not None:
                                # –ü—Ä–æ–±—É–µ–º —Ä–∞–∑–Ω—ã–µ —Å–ø–æ—Å–æ–±—ã –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞
                                try:
                                    # –°–Ω–∞—á–∞–ª–∞ –ø—Ä–æ–±—É–µ–º –∫–∞–∫ —á–∏—Å–ª–æ
                                    if isinstance(flight_number_raw, (int, float)):
                                        flight_number = int(float(flight_number_raw))
                                    else:
                                        # –ü—Ä–æ–±—É–µ–º –∏–∑–≤–ª–µ—á—å —á–∏—Å–ª–æ –∏–∑ —Å—Ç—Ä–æ–∫–∏
                                        flight_number_str = str(flight_number_raw).strip()
                                        # –£–±–∏—Ä–∞–µ–º –≤—Å–µ –Ω–µ—Ü–∏—Ñ—Ä–æ–≤—ã–µ —Å–∏–º–≤–æ–ª—ã, –∫—Ä–æ–º–µ –º–∏–Ω—É—Å–∞ –≤ –Ω–∞—á–∞–ª–µ
                                        import re
                                        numbers = re.findall(r'-?\d+', flight_number_str)
                                        if numbers:
                                            flight_number = int(float(numbers[0]))
                                except (ValueError, TypeError):
                                    # –ï—Å–ª–∏ –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π, –ø—Ä–æ—Å—Ç–æ –∏–≥–Ω–æ—Ä–∏—Ä—É–µ–º –µ–≥–æ
                                    flight_number = None
                            
                            # –ï—Å–ª–∏ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞ –Ω–µ—Ç - –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º –µ–≥–æ –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –Ω–∞ –æ—Å–Ω–æ–≤–µ –¥–∞—Ç—ã, –ø–∏–ª–æ—Ç–∞ –∏ –≤—Ä–µ–º–µ–Ω–∏
                            # –≠—Ç–æ –ø–æ–∑–≤–æ–ª–∏—Ç –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—Ç—å —Å—Ç—Ä–æ–∫–∏ –±–µ–∑ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞, –Ω–æ —Å –¥–∞—Ç–æ–π, –ø–æ–∑—ã–≤–Ω—ã–º –∏ —Ç–∏–ø–æ–º –¥—Ä–æ–Ω–∞
                            # –°–Ω–∞—á–∞–ª–∞ –ø–æ–ª—É—á–∞–µ–º –ø–∏–ª–æ—Ç–∞ –∏ –¥–∞—Ç—É, —á—Ç–æ–±—ã —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞—Ç—å –Ω–æ–º–µ—Ä

                            # –ü–æ–∑—ã–≤–Ω–æ–π –æ–ø–µ—Ä–∞—Ç–æ—Ä–∞ –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ V
                            pilot_callname_raw = row_values[COL_OPERATOR_CALLNAME - 1] if len(row_values) >= COL_OPERATOR_CALLNAME else None
                            pilot = None
                            if pilot_callname_raw:
                                callname_to_search = str(pilot_callname_raw).strip()
                                if callname_to_search.startswith("–ø–∏–ª–æ—Ç "):
                                    parts = callname_to_search.split()
                                    if len(parts) > 1:
                                        callname_to_search = parts[1]  # –ë–µ—Ä–µ–º –≤—Ç–æ—Ä—É—é —á–∞—Å—Ç—å
                                
                                # –ï—Å–ª–∏ –ø–æ–∑—ã–≤–Ω–æ–π –ø—É—Å—Ç–æ–π –ø–æ—Å–ª–µ –æ—á–∏—Å—Ç–∫–∏, —Å–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω–æ–≥–æ –ø–∏–ª–æ—Ç–∞
                                if not callname_to_search:
                                    callname_to_search = f"–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π_{row_idx}"
                                    logger.debug(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –ø—É—Å—Ç–æ–π –ø–æ–∑—ã–≤–Ω–æ–π, —Å–æ–∑–¥–∞–Ω –≤—Ä–µ–º–µ–Ω–Ω—ã–π: '{callname_to_search}'")
                                
                                # –ò—â–µ–º –ø–∏–ª–æ—Ç–∞ –≤ –∫—ç—à–µ (–±—ã—Å—Ç—Ä–µ–µ —á–µ–º –∑–∞–ø—Ä–æ—Å –∫ –ë–î)
                                callname_lower = callname_to_search.lower()
                                pilot = pilots_cache.get(callname_lower)
                                
                                if pilot is None:
                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º —Å–ø–∏—Å–æ–∫ –Ω–æ–≤—ã—Ö –ø–∏–ª–æ—Ç–æ–≤
                                    if callname_to_search in new_pilots:
                                        pilot = new_pilots[callname_to_search]
                                    else:
                                        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤–æ–≥–æ –ø–∏–ª–æ—Ç–∞ –∏ —Å–æ—Ö—Ä–∞–Ω—è–µ–º —Å—Ä–∞–∑—É –≤ –ë–î
                                        import uuid
                                        temp_tg_id = abs(hash(callname_to_search)) % (10 ** 10)
                                        while Pilot.objects.filter(tg_id=temp_tg_id).exists():
                                            temp_tg_id = abs(hash(f"{callname_to_search}{uuid.uuid4()}")) % (10 ** 10)
                                        
                                        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø–∏–ª–æ—Ç–∞ —Å—Ä–∞–∑—É, —á—Ç–æ–±—ã –ø–æ–ª—É—á–∏—Ç—å id –¥–ª—è foreign key
                                        pilot, created = Pilot.objects.get_or_create(
                                            callname=callname_to_search,
                                            defaults={'tg_id': temp_tg_id}
                                        )
                                        if created:
                                            logger.info(f"–ê–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Å–æ–∑–¥–∞–Ω –ø–∏–ª–æ—Ç '{callname_to_search}' —Å –≤—Ä–µ–º–µ–Ω–Ω—ã–º TG ID: {temp_tg_id}")
                                        
                                        # –î–æ–±–∞–≤–ª—è–µ–º –≤ –∫—ç—à–∏
                                        new_pilots[callname_to_search] = pilot
                                        pilots_cache[callname_lower] = pilot
                            else:
                                # –ï—Å–ª–∏ –ø–∏–ª–æ—Ç–∞ –Ω–µ—Ç, –Ω–æ –µ—Å—Ç—å –¥—Ä—É–≥–∏–µ –¥–∞–Ω–Ω—ã–µ - —Å–æ–∑–¥–∞–µ–º –≤—Ä–µ–º–µ–Ω–Ω–æ–≥–æ –ø–∏–ª–æ—Ç–∞
                                # –ò—Å–ø–æ–ª—å–∑—É–µ–º –Ω–æ–º–µ—Ä —Å—Ç—Ä–æ–∫–∏ –∫–∞–∫ –∏–¥–µ–Ω—Ç–∏—Ñ–∏–∫–∞—Ç–æ—Ä
                                import uuid
                                temp_callname = f"–ù–µ–∏–∑–≤–µ—Å—Ç–Ω—ã–π_{row_idx}"
                                temp_tg_id = abs(hash(f"temp_{row_idx}")) % (10 ** 10)
                                while Pilot.objects.filter(tg_id=temp_tg_id).exists():
                                    temp_tg_id = abs(hash(f"temp_{row_idx}_{uuid.uuid4()}")) % (10 ** 10)
                                
                                pilot, created = Pilot.objects.get_or_create(
                                    callname=temp_callname,
                                    defaults={'tg_id': temp_tg_id}
                                )
                                if created:
                                    logger.info(f"–°–æ–∑–¥–∞–Ω –≤—Ä–µ–º–µ–Ω–Ω—ã–π –ø–∏–ª–æ—Ç '{temp_callname}' –¥–ª—è —Å—Ç—Ä–æ–∫–∏ {row_idx}")
                                pilots_cache[temp_callname.lower()] = pilot

                            # –í—Ä–µ–º—è –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ B
                            time_str = row_values[COL_TIME - 1] if len(row_values) >= COL_TIME else None
                            flight_time = None
                            if time_str and isinstance(time_str, datetime.time):
                                flight_time = time_str
                            elif time_str and isinstance(time_str, str):
                                time_str_clean = str(time_str).strip()
                                if time_str_clean:
                                    try:
                                        flight_time = datetime.datetime.strptime(time_str_clean, "%H:%M").time()
                                    except ValueError:
                                        try:
                                            flight_time = datetime.datetime.strptime(time_str_clean, "%H:%M:%S").time()
                                        except ValueError:
                                            # –ü—Ä–æ–±—É–µ–º –¥—Ä—É–≥–∏–µ —Ñ–æ—Ä–º–∞—Ç—ã
                                            try:
                                                # –§–æ—Ä–º–∞—Ç "HH:MM:SS" –∏–ª–∏ "HH:MM"
                                                parts = time_str_clean.split(':')
                                                if len(parts) >= 2:
                                                    hour = int(parts[0])
                                                    minute = int(parts[1])
                                                    second = int(parts[2]) if len(parts) > 2 else 0
                                                    flight_time = datetime.time(hour, minute, second)
                                            except (ValueError, IndexError):
                                                pass
                            elif time_str and isinstance(time_str, datetime.datetime):
                                flight_time = time_str.time()
                            
                            # –ï—Å–ª–∏ –≤—Ä–µ–º—è –Ω–µ —É–¥–∞–ª–æ—Å—å —Ä–∞—Å–ø–∞—Ä—Å–∏—Ç—å, –∏—Å–ø–æ–ª—å–∑—É–µ–º –∑–Ω–∞—á–µ–Ω–∏–µ –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é (00:00:00)
                            if flight_time is None:
                                flight_time = datetime.time(0, 0, 0)

                            # –î–∞—Ç–∞ –≤—ã–ª–µ—Ç–∞ –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ R
                            flight_date = None
                            date_value = row_values[COL_FLIGHT_DATE - 1] if len(row_values) >= COL_FLIGHT_DATE else None
                            if date_value:
                                if isinstance(date_value, datetime.datetime):
                                    flight_date = date_value.date()
                                elif isinstance(date_value, datetime.date):
                                    flight_date = date_value
                                elif isinstance(date_value, str):
                                    # –ü—Ä–æ–±—É–µ–º —Ä–∞—Å–ø–∞—Ä—Å–∏—Ç—å —Å—Ç—Ä–æ–∫—É —Å –¥–∞—Ç–æ–π
                                    for date_format in ['%d.%m.%Y', '%d-%m-%Y', '%d/%m/%Y', '%Y-%m-%d', '%d.%m.%y', '%d-%m-%y']:
                                        try:
                                            flight_date = datetime.datetime.strptime(date_value.strip(), date_format).date()
                                            break
                                        except ValueError:
                                            continue
                            
                            # –ï—Å–ª–∏ –¥–∞—Ç—ã –Ω–µ—Ç - –∏—Å–ø–æ–ª—å–∑—É–µ–º —Ç–µ–∫—É—â—É—é –¥–∞—Ç—É
                            if flight_date is None:
                                flight_date = datetime.date.today()
                                logger.debug(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –Ω–µ—Ç –¥–∞—Ç—ã –≤—ã–ª–µ—Ç–∞, –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è —Ç–µ–∫—É—â–∞—è –¥–∞—Ç–∞ {flight_date}")
                            
                            # –ï—Å–ª–∏ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞ –Ω–µ—Ç, –Ω–æ –µ—Å—Ç—å –¥–∞—Ç–∞ –∏ –ø–∏–ª–æ—Ç - –≥–µ–Ω–µ—Ä–∏—Ä—É–µ–º –Ω–æ–º–µ—Ä –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏
                            # –ü–∏–ª–æ—Ç –≤—Å–µ–≥–¥–∞ –¥–æ–ª–∂–µ–Ω –±—ã—Ç—å —Å–æ–∑–¥–∞–Ω –∫ —ç—Ç–æ–º—É –º–æ–º–µ–Ω—Ç—É (–ª–∏–±–æ –∏–∑ –¥–∞–Ω–Ω—ã—Ö, –ª–∏–±–æ –≤—Ä–µ–º–µ–Ω–Ω—ã–π)
                            if flight_number is None:
                                if flight_date and pilot:
                                    # –ì–µ–Ω–µ—Ä–∏—Ä—É–µ–º –Ω–æ–º–µ—Ä –Ω–∞ –æ—Å–Ω–æ–≤–µ –¥–∞—Ç—ã, –ø–∏–ª–æ—Ç–∞ –∏ –Ω–æ–º–µ—Ä–∞ —Å—Ç—Ä–æ–∫–∏
                                    import hashlib
                                    unique_str = f"{flight_date.isoformat()}_{pilot.id}_{row_idx}"
                                    hash_value = int(hashlib.md5(unique_str.encode()).hexdigest()[:8], 16)
                                    flight_number = abs(hash_value) % (10 ** 8)
                                    # –£–±–µ–∂–¥–∞–µ–º—Å—è, —á—Ç–æ –Ω–æ–º–µ—Ä –Ω–µ –∫–æ–Ω—Ñ–ª–∏–∫—Ç—É–µ—Ç —Å —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–º–∏ –ø–æ–ª–µ—Ç–∞–º–∏ —ç—Ç–æ–≥–æ –ø–∏–ª–æ—Ç–∞
                                    while Flight.objects.filter(number=flight_number, pilot=pilot).exists():
                                        flight_number = (flight_number + 1) % (10 ** 8)
                                    logger.debug(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ {flight_number}")
                                else:
                                    # –ï—Å–ª–∏ –Ω–µ—Ç –¥–∞—Ç—ã –∏–ª–∏ –ø–∏–ª–æ—Ç–∞ - —ç—Ç–æ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞, –Ω–æ –ø–∏–ª–æ—Ç –¥–æ–ª–∂–µ–Ω –±—ã—Ç—å —Å–æ–∑–¥–∞–Ω
                                    # –ï—Å–ª–∏ –ø–∏–ª–æ—Ç–∞ –Ω–µ—Ç - —ç—Ç–æ –æ—à–∏–±–∫–∞ –≤ –ª–æ–≥–∏–∫–µ –≤—ã—à–µ
                                    if not pilot:
                                        logger.error(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –ö–†–ò–¢–ò–ß–ï–°–ö–ê–Ø –û–®–ò–ë–ö–ê - –ø–∏–ª–æ—Ç –Ω–µ —Å–æ–∑–¥–∞–Ω!")
                                        skipped_no_pilot += 1
                                        continue
                                    if not flight_date:
                                        logger.warning(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –Ω–µ—Ç –¥–∞—Ç—ã, –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è —Ç–µ–∫—É—â–∞—è –¥–∞—Ç–∞")
                                        flight_date = datetime.date.today()
                                    # –ü–æ–≤—Ç–æ—Ä—è–µ–º –≥–µ–Ω–µ—Ä–∞—Ü–∏—é –Ω–æ–º–µ—Ä–∞
                                    import hashlib
                                    unique_str = f"{flight_date.isoformat()}_{pilot.id}_{row_idx}"
                                    hash_value = int(hashlib.md5(unique_str.encode()).hexdigest()[:8], 16)
                                    flight_number = abs(hash_value) % (10 ** 8)
                                    while Flight.objects.filter(number=flight_number, pilot=pilot).exists():
                                        flight_number = (flight_number + 1) % (10 ** 8)
                                    logger.debug(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ {flight_number} (–ø–æ—Å–ª–µ –∏—Å–ø—Ä–∞–≤–ª–µ–Ω–∏—è)")
                            
                            # –¢–µ–ø–µ—Ä—å –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ –¥–æ–ª–∂–µ–Ω –±—ã—Ç—å –≤—Å–µ–≥–¥–∞ (–ª–∏–±–æ –∏–∑ –¥–∞–Ω–Ω—ã—Ö, –ª–∏–±–æ —Å–≥–µ–Ω–µ—Ä–∏—Ä–æ–≤–∞–Ω)
                            if flight_number is None:
                                # –≠—Ç–æ –Ω–µ –¥–æ–ª–∂–Ω–æ –ø—Ä–æ–∏—Å—Ö–æ–¥–∏—Ç—å, –Ω–æ –Ω–∞ –≤—Å—è–∫–∏–π —Å–ª—É—á–∞–π
                                skipped_no_flight_number += 1
                                if skipped_no_flight_number <= 20:
                                    logger.error(f"–°—Ç—Ä–æ–∫–∞ {row_idx}: –ö–†–ò–¢–ò–ß–ï–°–ö–ê–Ø –û–®–ò–ë–ö–ê - –Ω–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞! –î–∞—Ç–∞: {flight_date}, –ü–∏–ª–æ—Ç: {pilot.callname if pilot else None}")
                                continue
                            
                            # –õ–æ–≥–∏—Ä—É–µ–º –∫–∞–∂–¥—É—é 1000-—é —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—É—é —Å—Ç—Ä–æ–∫—É –¥–ª—è –æ—Ç–ª–∞–¥–∫–∏
                            if processed_successfully > 0 and processed_successfully % 1000 == 0:
                                logger.info(f"–£—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ —Å—Ç—Ä–æ–∫: {processed_successfully}, —Å–æ–∑–¥–∞–Ω–æ –∑–∞–ø–∏—Å–µ–π –≤ –ø–∞–º—è—Ç–∏: {len(flights_to_create)}, –≤—Å–µ–≥–æ —Å–æ–∑–¥–∞–Ω–æ –≤ –ë–î: {created_count}")

                            # –•–∞—Ä–∞–∫—Ç–µ—Ä —Ü–µ–ª–∏ –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ C
                            target_raw = row_values[COL_TARGET - 1] if len(row_values) >= COL_TARGET else None
                            target_raw = str(target_raw).strip() if target_raw else None
                            
                            # –ù–æ—Ä–º–∞–ª–∏–∑—É–µ–º –Ω–∞–∑–≤–∞–Ω–∏–µ —Ü–µ–ª–∏ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –¥—É–±–ª–∏—Ä–æ–≤–∞–Ω–∏—è
                            target = normalize_target_name(target_raw) if target_raw else None
                            
                            # –î–æ–±–∞–≤–ª—è–µ–º –≤ —Å–ø–∏—Å–æ–∫ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è, –µ—Å–ª–∏ –Ω—É–∂–Ω–æ (—Å–æ–∑–¥–∞–¥–∏–º –±–∞—Ç—á–µ–º –ø–æ–∑–∂–µ)
                            if target:
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –∫—ç—à–µ –ø–æ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–º—É –∑–Ω–∞—á–µ–Ω–∏—é
                                target_lower = target.lower()
                                if target_lower not in target_types_cache:
                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ —É–∂–µ –≤ —Å–ø–∏—Å–∫–µ –Ω–æ–≤—ã—Ö —Å —Ç–∞–∫–∏–º –∂–µ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–º –∑–Ω–∞—á–µ–Ω–∏–µ–º
                                    existing_in_new = None
                                    for existing_target in new_target_types.keys():
                                        if existing_target.lower() == target_lower:
                                            existing_in_new = existing_target
                                            break
                                    
                                    if existing_in_new:
                                        # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–µ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
                                        target = existing_in_new
                                    else:
                                        # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤–æ–µ —Å –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã–º –∏–º–µ–Ω–µ–º
                                        new_target_types[target] = TargetType(name=target, weight=1)
                                else:
                                    # –ò—Å–ø–æ–ª—å–∑—É–µ–º –∑–Ω–∞—á–µ–Ω–∏–µ –∏–∑ –∫—ç—à–∞ (—É–∂–µ –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω–æ–µ)
                                    target = target_types_cache[target_lower].name
                            
                            corrective = None  # –í –Ω–æ–≤–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä–µ –Ω–µ—Ç –æ—Ç–¥–µ–ª—å–Ω–æ–π –∫–æ–ª–æ–Ω–∫–∏ –¥–ª—è corrective

                            # –†–µ–∑—É–ª—å—Ç–∞—Ç –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ O
                            result_raw = row_values[COL_RESULT - 1] if len(row_values) >= COL_RESULT else None
                            result = FlightResultTypes.NOT_DEFEATED
                            if result_raw:
                                result_str = str(result_raw).lower().strip()
                                if "—É–Ω–∏—á—Ç–æ–∂" in result_str:
                                    result = FlightResultTypes.DESTROYED
                                elif "–Ω–µ –ø" in result_str:
                                    result = FlightResultTypes.NOT_DEFEATED
                                elif "–ø–æ–¥–∞–≤–ª" in result_str or "—É—Å–ø–µ—à" in result_str or "–ø–æ—Ä–∞–∂" in result_str:
                                    result = FlightResultTypes.DEFEATED

                            # –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏–∑ –∫–æ–ª–æ–Ω–æ–∫ E –∏ F
                            coord_x = row_values[COL_COORD_X - 1] if len(row_values) >= COL_COORD_X else None
                            coord_y = row_values[COL_COORD_Y - 1] if len(row_values) >= COL_COORD_Y else None
                            coordinates = None
                            if coord_x is not None and coord_y is not None:
                                coordinates = f"{coord_x} {coord_y}"

                            # –¢–∏–ø –¥—Ä–æ–Ω–∞ –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ H
                            drone_raw = row_values[COL_DRONE - 1] if len(row_values) >= COL_DRONE else None
                            drone_raw = str(drone_raw).strip() if drone_raw else None
                            
                            # –ù–æ—Ä–º–∞–ª–∏–∑—É–µ–º –Ω–∞–∑–≤–∞–Ω–∏–µ –¥—Ä–æ–Ω–∞ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –¥—É–±–ª–∏—Ä–æ–≤–∞–Ω–∏—è
                            drone = normalize_drone_name(drone_raw) if drone_raw else None
                            
                            # –î–æ–±–∞–≤–ª—è–µ–º –≤ —Å–ø–∏—Å–æ–∫ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è, –µ—Å–ª–∏ –Ω—É–∂–Ω–æ (—Å–æ–∑–¥–∞–¥–∏–º –±–∞—Ç—á–µ–º –ø–æ–∑–∂–µ)
                            if drone:
                                # –°–æ–∑–¥–∞–µ–º –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤ (—á—Ç–æ–±—ã X-51 –∏ X51 –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∞–ª–∏—Å—å)
                                def get_drone_comparison_key(drone_name):
                                    """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –¥—Ä–æ–Ω–æ–≤ (–±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤)"""
                                    return re.sub(r'[-]', '', drone_name.lower().strip())
                                
                                drone_key = get_drone_comparison_key(drone)
                                
                                # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –∫—ç—à–µ –ø–æ –∫–ª—é—á—É –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤
                                found_in_cache = None
                                for cached_key, cached_drone in drones_cache.items():
                                    if get_drone_comparison_key(cached_drone.name) == drone_key:
                                        found_in_cache = cached_drone
                                        # –ï—Å–ª–∏ –Ω–æ–≤—ã–π –¥—Ä–æ–Ω –∏–º–µ–µ—Ç –¥–µ—Ñ–∏—Å, –∞ –Ω–∞–π–¥–µ–Ω–Ω—ã–π - –Ω–µ—Ç, –æ–±–Ω–æ–≤–ª—è–µ–º
                                        if '-' in drone and '-' not in cached_drone.name:
                                            cached_drone.name = drone
                                            cached_drone.save(update_fields=['name'])
                                        drone = found_in_cache.name
                                        break
                                
                                if not found_in_cache:
                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –Ω–µ—Ç –ª–∏ —É–∂–µ –≤ —Å–ø–∏—Å–∫–µ –Ω–æ–≤—ã—Ö —Å —Ç–∞–∫–∏–º –∂–µ –∫–ª—é—á–æ–º
                                    existing_in_new = None
                                    for existing_drone in new_drones.keys():
                                        if get_drone_comparison_key(existing_drone) == drone_key:
                                            existing_in_new = existing_drone
                                            # –ï—Å–ª–∏ –Ω–æ–≤—ã–π –¥—Ä–æ–Ω –∏–º–µ–µ—Ç –¥–µ—Ñ–∏—Å, –∏—Å–ø–æ–ª—å–∑—É–µ–º –µ–≥–æ
                                            if '-' in drone and '-' not in existing_drone:
                                                # –û–±–Ω–æ–≤–ª—è–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–∏–π –≤ new_drones
                                                drone_obj = new_drones.pop(existing_drone)
                                                drone_obj.name = drone
                                                new_drones[drone] = drone_obj
                                                existing_in_new = drone
                                            break
                                    
                                    if existing_in_new:
                                        # –ò—Å–ø–æ–ª—å–∑—É–µ–º —Å—É—â–µ—Å—Ç–≤—É—é—â–µ–µ –∑–Ω–∞—á–µ–Ω–∏–µ
                                        drone = existing_in_new
                                    else:
                                        # –û–ø—Ä–µ–¥–µ–ª—è–µ–º —Ç–∏–ø –¥—Ä–æ–Ω–∞ (KT –∏–ª–∏ ST) –ø–æ –Ω–∞–∑–≤–∞–Ω–∏—é
                                        drone_lower = drone.lower()
                                        drone_type_choice = DroneTypes.KT  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é KT
                                        if '—Å—Ç' in drone_lower or 'st' in drone_lower:
                                            drone_type_choice = DroneTypes.ST
                                        elif '–∫—Ç' in drone_lower or 'kt' in drone_lower:
                                            drone_type_choice = DroneTypes.KT
                                        new_drones[drone] = Drone(name=drone, drone_type=drone_type_choice, description='')
                            
                            # –í–∏–¥ –ë–ü –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ J
                            explosive_type_raw = row_values[COL_EXPLOSIVE_TYPE - 1] if len(row_values) >= COL_EXPLOSIVE_TYPE else None
                            explosive_type = None
                            if explosive_type_raw:
                                explosive_type_str = str(explosive_type_raw).strip()
                                if explosive_type_str:
                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –∫—ç—à–µ
                                    explosive_type_lower = explosive_type_str.lower()
                                    explosive_type_obj = explosive_types_cache.get(explosive_type_lower)
                                    
                                    if explosive_type_obj is None:
                                        # –î–æ–±–∞–≤–ª—è–µ–º –≤ —Å–ø–∏—Å–æ–∫ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è
                                        if explosive_type_str not in new_explosive_types:
                                            new_explosive_types[explosive_type_str] = ExplosiveType(name=explosive_type_str)
                                        explosive_type = explosive_type_str
                                    else:
                                        explosive_type = explosive_type_obj.name
                            
                            # –í–∏–¥ –≤–∑—Ä—ã–≤–∞—Ç–µ–ª—è –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ L
                            explosive_device_raw = row_values[COL_EXPLOSIVE_DEVICE - 1] if len(row_values) >= COL_EXPLOSIVE_DEVICE else None
                            explosive_device = None
                            if explosive_device_raw:
                                explosive_device_str = str(explosive_device_raw).strip()
                                if explosive_device_str:
                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –∫—ç—à–µ
                                    explosive_device_lower = explosive_device_str.lower()
                                    explosive_device_obj = explosive_devices_cache.get(explosive_device_lower)
                                    
                                    if explosive_device_obj is None:
                                        # –î–æ–±–∞–≤–ª—è–µ–º –≤ —Å–ø–∏—Å–æ–∫ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è
                                        if explosive_device_str not in new_explosive_devices:
                                            new_explosive_devices[explosive_device_str] = ExplosiveDevice(name=explosive_device_str)
                                        explosive_device = explosive_device_str
                                    else:
                                        explosive_device = explosive_device_obj.name
                            
                            # –î–∏—Å—Ç–∞–Ω—Ü–∏—è –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ T
                            distance = row_values[COL_DISTANCE - 1] if len(row_values) >= COL_DISTANCE else None
                            distance = str(distance).strip() if distance else None
                            
                            # –ü–æ–ª—è, –∫–æ—Ç–æ—Ä—ã—Ö –Ω–µ—Ç –≤ –Ω–æ–≤–æ–π —Å—Ç—Ä—É–∫—Ç—É—Ä–µ, –æ—Å—Ç–∞–≤–ª—è–µ–º None
                            engineer = None
                            driver = None
                            video = None
                            manage = None
                            video_length = None
                            
                            # –ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π –∏–∑ –∫–æ–ª–æ–Ω–∫–∏ D
                            comment_raw = row_values[COL_COMMENT - 1] if len(row_values) >= COL_COMMENT else None
                            comment = None
                            if comment_raw:
                                try:
                                    comment = str(comment_raw).strip()
                                    if comment and comment.lower() not in ('none', 'null', ''):
                                        # –û–±—Ä–µ–∑–∞–µ–º –¥–æ 255 —Å–∏–º–≤–æ–ª–æ–≤ (–æ–≥—Ä–∞–Ω–∏—á–µ–Ω–∏–µ –ø–æ–ª—è –≤ –ë–î)
                                        if len(comment) > 255:
                                            comment = comment[:255]
                                    else:
                                        comment = None
                                except Exception as comment_error:
                                    logger.warning(f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏—è –≤ —Å—Ç—Ä–æ–∫–µ {row_idx}: {comment_error}")
                                    comment = None
                            
                            drone_remains = None
                            direction = None
                            objective = FlightObjectiveTypes.NOT_EXISTS  # –ü–æ —É–º–æ–ª—á–∞–Ω–∏—é
                            
                            # –ü–æ–¥–≥–æ—Ç–∞–≤–ª–∏–≤–∞–µ–º –¥–∞–Ω–Ω—ã–µ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è
                            flight_defaults = {
                                'flight_date': flight_date,
                                'flight_time': flight_time,
                                'target': target,
                                'corrective': corrective,
                                'result': result,
                                'drone': drone,
                                'explosive_type': explosive_type,
                                'explosive_device': explosive_device,
                                'distance': distance,
                                'objective': objective,
                                'comment': comment,
                            }
                            
                            # –î–æ–±–∞–≤–ª—è–µ–º coordinates —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –æ–Ω–∏ –µ—Å—Ç—å
                            if coordinates and coordinates.strip():
                                flight_defaults['coordinates'] = coordinates
                            
                            # –£–¥–∞–ª—è–µ–º None –∑–Ω–∞—á–µ–Ω–∏—è, –Ω–æ –æ—Å—Ç–∞–≤–ª—è–µ–º comment –¥–∞–∂–µ –µ—Å–ª–∏ –æ–Ω None (—ç—Ç–æ –≤–∞–ª–∏–¥–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ)
                            # comment –º–æ–∂–µ—Ç –±—ã—Ç—å None, –∏ —ç—Ç–æ –Ω–æ—Ä–º–∞–ª—å–Ω–æ
                            flight_defaults_clean = {k: v for k, v in flight_defaults.items() if v is not None or k == 'comment'}
                            flight_defaults = flight_defaults_clean
                            
                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –ª–∏ —É–∂–µ —Ç–∞–∫–æ–π –ø–æ–ª–µ—Ç
                            # –°–Ω–∞—á–∞–ª–∞ –ø—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ –∫–ª—é—á–µ–≤—ã–º –ø–æ–ª—è–º (–±—ã—Å—Ç—Ä–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞)
                            flight_key = (
                                flight_number,
                                pilot.id if pilot else None,
                                flight_defaults.get('flight_date'),
                                flight_defaults.get('flight_time')
                            )
                            
                            # –ü—Ä–æ–ø—É—Å–∫–∞–µ–º, –µ—Å–ª–∏ –ø–æ–ª–µ—Ç —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –≤ —Ç–µ–∫—É—â–µ–º –∏–º–ø–æ—Ä—Ç–µ
                            if flight_key in existing_flights_set:
                                skipped_duplicates += 1
                                if skipped_duplicates <= 10:  # –õ–æ–≥–∏—Ä—É–µ–º —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 10
                                    logger.debug(f"–ü—Ä–æ–ø—É—â–µ–Ω –¥—É–±–ª–∏–∫–∞—Ç –ø–æ–ª–µ—Ç–∞ –≤ —Ç–µ–∫—É—â–µ–º –∏–º–ø–æ—Ä—Ç–µ: –Ω–æ–º–µ—Ä={flight_number}, –ø–∏–ª–æ—Ç={pilot.callname if pilot else None}, –¥–∞—Ç–∞={flight_defaults.get('flight_date')}, –≤—Ä–µ–º—è={flight_defaults.get('flight_time')}")
                                continue
                            
                            # –°—Ç—Ä–æ–≥–∞—è –ø—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –ø–æ–ª–Ω–æ—Å—Ç—å—é –æ–¥–∏–Ω–∞–∫–æ–≤—ã–µ –∑–∞–ø–∏—Å–∏ (–≤—Å–µ –ø–æ–ª—è)
                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –≤ –±–∞–∑–µ –¥–∞–Ω–Ω—ã—Ö –Ω–∞ –ø–æ–ª–Ω–æ—Å—Ç—å—é –æ–¥–∏–Ω–∞–∫–æ–≤—ã–µ –∑–∞–ø–∏—Å–∏ —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –Ω–µ –Ω–∞–π–¥–µ–Ω –≤ —Ç–µ–∫—É—â–µ–º –∏–º–ø–æ—Ä—Ç–µ
                            # –≠—Ç–æ –æ–ø—Ç–∏–º–∏–∑–∞—Ü–∏—è - –∏–∑–±–µ–≥–∞–µ–º –ª–∏—à–Ω–∏—Ö –∑–∞–ø—Ä–æ—Å–æ–≤ –∫ –ë–î
                            existing_duplicate = None
                            try:
                                existing_duplicate = Flight.objects.filter(
                                    number=flight_number,
                                    pilot=pilot,
                                    flight_date=flight_defaults.get('flight_date'),
                                    flight_time=flight_defaults.get('flight_time'),
                                    target=flight_defaults.get('target'),
                                    drone=flight_defaults.get('drone'),
                                    result=flight_defaults.get('result'),
                                    coordinates=flight_defaults.get('coordinates'),
                                    distance=flight_defaults.get('distance'),
                                    explosive_type=flight_defaults.get('explosive_type'),
                                    explosive_device=flight_defaults.get('explosive_device'),
                                    corrective=flight_defaults.get('corrective'),
                                    objective=flight_defaults.get('objective'),
                                    comment=flight_defaults.get('comment'),
                                ).first()
                            except Exception as dup_check_error:
                                logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –ø—Ä–æ–≤–µ—Ä–∫–µ –¥—É–±–ª–∏–∫–∞—Ç–∞ –¥–ª—è —Å—Ç—Ä–æ–∫–∏ {row_idx}: {dup_check_error}")
                            
                            if existing_duplicate:
                                skipped_duplicates += 1
                                if skipped_duplicates <= 10:  # –õ–æ–≥–∏—Ä—É–µ–º —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 10
                                    logger.debug(f"–ü—Ä–æ–ø—É—â–µ–Ω –ø–æ–ª–Ω–æ—Å—Ç—å—é –∏–¥–µ–Ω—Ç–∏—á–Ω—ã–π –¥—É–±–ª–∏–∫–∞—Ç –ø–æ–ª–µ—Ç–∞: –Ω–æ–º–µ—Ä={flight_number}, –ø–∏–ª–æ—Ç={pilot.callname if pilot else None}, –¥–∞—Ç–∞={flight_defaults.get('flight_date')}, –≤—Ä–µ–º—è={flight_defaults.get('flight_time')}")
                                continue
                            
                            # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—ã–π –ø–æ–ª–µ—Ç
                            try:
                                new_flight = Flight(
                                    number=flight_number,
                                    pilot=pilot,
                                    **flight_defaults
                                )
                                flights_to_create.append(new_flight)
                                # –î–æ–±–∞–≤–ª—è–µ–º –≤ –º–Ω–æ–∂–µ—Å—Ç–≤–æ —Å—É—â–µ—Å—Ç–≤—É—é—â–∏—Ö, —á—Ç–æ–±—ã –Ω–µ —Å–æ–∑–¥–∞–≤–∞—Ç—å –¥—É–±–ª–∏–∫–∞—Ç—ã –≤ —Ä–∞–º–∫–∞—Ö –æ–¥–Ω–æ–≥–æ –∏–º–ø–æ—Ä—Ç–∞
                                existing_flights_set.add(flight_key)
                                processed_successfully += 1
                                
                                # –õ–æ–≥–∏—Ä—É–µ–º –∫–∞–∂–¥—É—é 1000-—é —É—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–Ω—É—é —Å—Ç—Ä–æ–∫—É
                                if processed_successfully % 1000 == 0:
                                    logger.info(f"–û–±—Ä–∞–±–æ—Ç–∞–Ω–æ —É—Å–ø–µ—à–Ω–æ: {processed_successfully}, –≤ –ø–∞–º—è—Ç–∏ –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è: {len(flights_to_create)}, –≤—Å–µ–≥–æ —Å–æ–∑–¥–∞–Ω–æ –≤ –ë–î: {created_count}")
                            except Exception as flight_create_error:
                                logger.error(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è –æ–±—ä–µ–∫—Ç–∞ Flight –¥–ª—è —Å—Ç—Ä–æ–∫–∏ {row_idx}: {flight_create_error}")
                                logger.error(f"–î–∞–Ω–Ω—ã–µ: –Ω–æ–º–µ—Ä={flight_number}, –ø–∏–ª–æ—Ç={pilot.callname if pilot else None}, defaults={flight_defaults}")
                                skipped_errors += 1
                                continue
                            
                            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –±–∞—Ç—á–∞–º–∏ –∫–∞–∂–¥—ã–µ 25 –∑–∞–ø–∏—Å–µ–π (—É–º–µ–Ω—å—à–µ–Ω–æ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –Ω–µ—Ö–≤–∞—Ç–∫–∏ –ø–∞–º—è—Ç–∏)
                            total_batch = len(flights_to_create)
                            if total_batch >= 25:
                                from django.db import transaction
                                try:
                                    with transaction.atomic():
                                        # –°–Ω–∞—á–∞–ª–∞ —Å–æ–∑–¥–∞–µ–º —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∏, –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
                                        if new_target_types:
                                            TargetType.objects.bulk_create(new_target_types.values(), ignore_conflicts=True)
                                            # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                                            target_types_cache = {tt.name.lower(): tt for tt in TargetType.objects.all()}
                                            new_target_types.clear()
                                        
                                        if new_drones:
                                            Drone.objects.bulk_create(new_drones.values(), ignore_conflicts=True)
                                            # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                                            # –°–æ–∑–¥–∞–µ–º –∫—ç—à –¥—Ä–æ–Ω–æ–≤ —Å –∫–ª—é—á–æ–º –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤ –¥–ª—è –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∏ X-51 –∏ X51
                                            def get_drone_comparison_key(drone_name):
                                                """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –¥—Ä–æ–Ω–æ–≤ (–±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤)"""
                                                return re.sub(r'[-]', '', str(drone_name).lower().strip())
                                            drones_cache = {get_drone_comparison_key(drone.name): drone for drone in Drone.objects.all()}
                                            new_drones.clear()
                                        
                                        if new_explosive_types:
                                            ExplosiveType.objects.bulk_create(new_explosive_types.values(), ignore_conflicts=True)
                                            for et in new_explosive_types.values():
                                                explosive_types_cache[et.name.lower()] = et
                                            new_explosive_types.clear()
                                        
                                        if new_explosive_devices:
                                            ExplosiveDevice.objects.bulk_create(new_explosive_devices.values(), ignore_conflicts=True)
                                            for ed in new_explosive_devices.values():
                                                explosive_devices_cache[ed.name.lower()] = ed
                                            new_explosive_devices.clear()
                                        
                                        if flights_to_create:
                                            # –°–æ—Ö—Ä–∞–Ω—è–µ–º –∫–æ–ø–∏—é –¥–ª—è –æ–±–Ω–æ–≤–ª–µ–Ω–∏—è existing_flights_set
                                            temp_flights_for_set = flights_to_create.copy()
                                            # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ –º–µ–Ω—å—à–∏–µ –±–∞—Ç—á–∏ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –Ω–µ—Ö–≤–∞—Ç–∫–∏ –ø–∞–º—è—Ç–∏
                                            batch_size = 25
                                            total_created_in_batch = 0
                                            total_failed_in_batch = 0
                                            for i in range(0, len(flights_to_create), batch_size):
                                                batch = flights_to_create[i:i + batch_size]
                                                try:
                                                    created_objects = Flight.objects.bulk_create(batch, ignore_conflicts=True)
                                                    # bulk_create —Å ignore_conflicts=True –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç –¢–û–õ–¨–ö–û —Å–æ–∑–¥–∞–Ω–Ω—ã–µ –æ–±—ä–µ–∫—Ç—ã
                                                    # –ï—Å–ª–∏ –æ–±—ä–µ–∫—Ç –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω –∏–∑-–∑–∞ –∫–æ–Ω—Ñ–ª–∏–∫—Ç–∞, –æ–Ω –ù–ï –±—É–¥–µ—Ç –≤ —Å–ø–∏—Å–∫–µ
                                                    created_count_in_sub_batch = len(created_objects) if created_objects else 0
                                                    total_created_in_batch += created_count_in_sub_batch
                                                    
                                                    # –ï—Å–ª–∏ —Å–æ–∑–¥–∞–Ω–æ –º–µ–Ω—å—à–µ, —á–µ–º –≤ –±–∞—Ç—á–µ - –∑–Ω–∞—á–∏—Ç –±—ã–ª–∏ –∫–æ–Ω—Ñ–ª–∏–∫—Ç—ã
                                                    if created_count_in_sub_batch < len(batch):
                                                        failed_count = len(batch) - created_count_in_sub_batch
                                                        total_failed_in_batch += failed_count
                                                        if total_failed_in_batch <= 10:  # –õ–æ–≥–∏—Ä—É–µ–º —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 10
                                                            logger.warning(f"–í –±–∞—Ç—á–µ {i//batch_size + 1} –Ω–µ —Å–æ–∑–¥–∞–Ω–æ {failed_count} –∑–∞–ø–∏—Å–µ–π –∏–∑-–∑–∞ –∫–æ–Ω—Ñ–ª–∏–∫—Ç–æ–≤ (—Å—Ç—Ä–æ–∫–∞ ~{row_idx})")
                                                        # –ü—Ä–æ–±—É–µ–º —Å–æ–∑–¥–∞—Ç—å –ø–æ –æ–¥–Ω–æ–π –∑–∞–ø–∏—Å–∏ –¥–ª—è —Ç–µ—Ö, —á—Ç–æ –Ω–µ —Å–æ–∑–¥–∞–ª–∏—Å—å
                                                        for flight in batch:
                                                            # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –±—ã–ª –ª–∏ —Å–æ–∑–¥–∞–Ω —ç—Ç–æ—Ç –ø–æ–ª–µ—Ç (–ø—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ –≤—Å–µ–º –∫–ª—é—á–µ–≤—ã–º –ø–æ–ª—è–º)
                                                            existing = Flight.objects.filter(
                                                                number=flight.number,
                                                                pilot=flight.pilot,
                                                                flight_date=flight.flight_date,
                                                                flight_time=flight.flight_time
                                                            ).first()
                                                            
                                                            if not existing:
                                                                try:
                                                                    flight.save()
                                                                    total_created_in_batch += 1
                                                                    total_failed_in_batch -= 1
                                                                    if total_failed_in_batch <= 5:
                                                                        logger.info(f"–£—Å–ø–µ—à–Ω–æ —Å–æ–∑–¥–∞–Ω –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname} –ø—Ä–∏ –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –ø–æ–ø—ã—Ç–∫–µ")
                                                                except Exception as single_error:
                                                                    if total_failed_in_batch <= 10:
                                                                        logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname}: {single_error}")
                                                            else:
                                                                # –ó–∞–ø–∏—Å—å —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç - —ç—Ç–æ –Ω–æ—Ä–º–∞–ª—å–Ω–æ, –ø—Ä–æ—Å—Ç–æ –Ω–µ —Å—á–∏—Ç–∞–µ–º –∫–∞–∫ –æ—à–∏–±–∫—É
                                                                total_failed_in_batch -= 1
                                                                if total_failed_in_batch <= 5:
                                                                    logger.debug(f"–ü–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname} —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –≤ –ë–î")
                                                except Exception as batch_create_error:
                                                    logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ bulk_create –±–∞—Ç—á–∞ {i//batch_size + 1}: {batch_create_error}", exc_info=True)
                                                    # –ü—Ä–æ–±—É–µ–º —Å–æ–∑–¥–∞—Ç—å –ø–æ –æ–¥–Ω–æ–π –∑–∞–ø–∏—Å–∏
                                                    for flight in batch:
                                                        try:
                                                            flight.save()
                                                            total_created_in_batch += 1
                                                        except Exception as single_error:
                                                            total_failed_in_batch += 1
                                                            if total_failed_in_batch <= 10:
                                                                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname}: {single_error}")
                                            
                                            created_count += total_created_in_batch
                                            if total_created_in_batch != len(flights_to_create):
                                                logger.warning(f"–°–æ–∑–¥–∞–Ω–æ {total_created_in_batch} –∏–∑ {len(flights_to_create)} –∑–∞–ø–∏—Å–µ–π –≤ –±–∞—Ç—á–µ (—Å—Ç—Ä–æ–∫–∞ ~{row_idx}). –ù–µ —Å–æ–∑–¥–∞–Ω–æ: {total_failed_in_batch}")
                                                self.message_user(request,
                                                                  _(f"‚ö†Ô∏è –í –±–∞—Ç—á–µ —Å–æ–∑–¥–∞–Ω–æ {total_created_in_batch} –∏–∑ {len(flights_to_create)} –∑–∞–ø–∏—Å–µ–π (—Å—Ç—Ä–æ–∫–∞ ~{row_idx}). –ù–µ —Å–æ–∑–¥–∞–Ω–æ: {total_failed_in_batch}"),
                                                                  level=messages.WARNING)
                                            
                                            # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –≤ —Ñ–æ–Ω–æ–≤–æ–º —Ä–µ–∂–∏–º–µ –¥–ª—è —Å–æ–∑–¥–∞–Ω–Ω—ã—Ö –ø–æ–ª–µ—Ç–æ–≤
                                            # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –≤—Å–µ –ø–æ–ª–µ—Ç—ã —Å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏, –∫–æ—Ç–æ—Ä—ã–µ –µ—â–µ –Ω–µ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω—ã
                                            import threading
                                            def convert_coordinates_background():
                                                """–§–æ–Ω–æ–≤–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç"""
                                                try:
                                                    from flights.models import Flight
                                                    # –ù–∞—Ö–æ–¥–∏–º –≤—Å–µ –ø–æ–ª–µ—Ç—ã —Å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏, –∫–æ—Ç–æ—Ä—ã–µ –µ—â–µ –Ω–µ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω—ã
                                                    flights_to_convert = Flight.objects.filter(
                                                        coordinates__isnull=False
                                                    ).exclude(
                                                        coordinates=''
                                                    ).filter(
                                                        lat_wgs84__isnull=True
                                                    )[:100]  # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –ø–æ 100 –∑–∞ —Ä–∞–∑
                                                    
                                                    converted_count = 0
                                                    for flight in flights_to_convert:
                                                        try:
                                                            # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏ —Å–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –ë–î
                                                            flight.get_coordinates_info_cached()
                                                            converted_count += 1
                                                        except Exception as coord_error:
                                                            logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –ø–æ–ª–µ—Ç–∞ {flight.id}: {coord_error}")
                                                    
                                                    if converted_count > 0:
                                                        logger.info(f"–ü—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –≤ —Ñ–æ–Ω–æ–≤–æ–º —Ä–µ–∂–∏–º–µ: {converted_count} –ø–æ–ª–µ—Ç–æ–≤")
                                                except Exception as bg_error:
                                                    logger.error(f"–û—à–∏–±–∫–∞ –≤ —Ñ–æ–Ω–æ–≤–æ–º –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç: {bg_error}", exc_info=True)
                                            
                                            # –ó–∞–ø—É—Å–∫–∞–µ–º –≤ —Ñ–æ–Ω–æ–≤–æ–º –ø–æ—Ç–æ–∫–µ
                                            thread = threading.Thread(target=convert_coordinates_background, daemon=True)
                                            thread.start()
                                            logger.debug(f"–ó–∞–ø—É—â–µ–Ω–∞ —Ñ–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç")
                                            
                                            # –û—á–∏—â–∞–µ–º —Å–ø–∏—Å–æ–∫ –ø–æ—Å–ª–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è
                                            # (–Ω–µ –¥–æ–±–∞–≤–ª—è–µ–º –≤ existing_flights_set, —Ç–∞–∫ –∫–∞–∫ –±–∞–∑–∞ —É–∂–µ –æ—á–∏—â–µ–Ω–∞)
                                            flights_to_create = []
                                        
                                        # –Ø–≤–Ω–∞—è –æ—á–∏—Å—Ç–∫–∞ –ø–∞–º—è—Ç–∏ –ø–æ—Å–ª–µ –∫–∞–∂–¥–æ–≥–æ –±–∞—Ç—á–∞
                                        import gc
                                        gc.collect()
                                        
                                        # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ –ø–æ—Å–ª–µ –∫–∞–∂–¥–æ–≥–æ –±–∞—Ç—á–∞
                                        if import_progress:
                                            try:
                                                import_progress.last_processed_row = row_idx
                                                import_progress.total_created = created_count
                                                import_progress.save(update_fields=['last_processed_row', 'total_created', 'last_import_date'])
                                                logger.debug(f"–ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω: —Å—Ç—Ä–æ–∫–∞ {row_idx}, —Å–æ–∑–¥–∞–Ω–æ {created_count}")
                                            except Exception as progress_error:
                                                logger.warning(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∏–º–ø–æ—Ä—Ç–∞: {progress_error}")
                                        
                                        # flights_to_update –Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è, —Ç–∞–∫ –∫–∞–∫ –±–∞–∑–∞ –æ—á–∏—â–µ–Ω–∞
                                    
                                    # –ö—ç—à –±–æ–ª—å—à–µ –Ω–µ –∏—Å–ø–æ–ª—å–∑—É–µ—Ç—Å—è, —Ç–∞–∫ –∫–∞–∫ –Ω–µ –ø—Ä–æ–≤–µ—Ä—è–µ–º –¥—É–±–ª–∏–∫–∞—Ç—ã
                                    
                                    # –ü–æ–∫–∞–∑—ã–≤–∞–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∫–∞–∂–¥—ã–µ 1000 —Å–æ—Ö—Ä–∞–Ω–µ–Ω–Ω—ã—Ö –∑–∞–ø–∏—Å–µ–π
                                    total_processed = created_count
                                    if total_processed % 1000 == 0:
                                        self.message_user(request,
                                                          _(f"üíæ –°–æ—Ö—Ä–∞–Ω–µ–Ω–æ {total_processed} –∑–∞–ø–∏—Å–µ–π (—Å–æ–∑–¥–∞–Ω–æ: {created_count})..."),
                                                          level=messages.INFO)
                                    
                                except Exception as batch_error:
                                    error_msg = f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –±–∞—Ç—á–∞ (—Å—Ç—Ä–æ–∫–∞ ~{row_idx}): {str(batch_error)}"
                                    if len(error_messages) < 200:
                                        error_messages.append(error_msg)
                                    logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –±–∞—Ç—á–∞: {batch_error}", exc_info=True)
                                    # –ù–ï –æ—á–∏—â–∞–µ–º —Å–ø–∏—Å–∫–∏ - –æ—Å—Ç–∞–≤–∏–º –∑–∞–ø–∏—Å–∏ –¥–ª—è –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –ø–æ–ø—ã—Ç–∫–∏ –≤ —Ñ–∏–Ω–∞–ª—å–Ω–æ–º –±–∞—Ç—á–µ
                                    # –¢–æ–ª—å–∫–æ –ª–æ–≥–∏—Ä—É–µ–º –æ—à–∏–±–∫—É –∏ –ø—Ä–æ–¥–æ–ª–∂–∞–µ–º
                                    self.message_user(request,
                                                      _(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –±–∞—Ç—á–∞ –Ω–∞ —Å—Ç—Ä–æ–∫–µ ~{row_idx}. –ó–∞–ø–∏—Å–∏ –±—É–¥—É—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω—ã –≤ —Ñ–∏–Ω–∞–ª—å–Ω–æ–º –±–∞—Ç—á–µ."),
                                                      level=messages.WARNING)

                        except Exception as e:
                            error_msg = f"–û—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å—Ç—Ä–æ–∫–∏ {row_idx} –≤ —Ñ–∞–π–ª–µ '{xlsx_file.name}': {str(e)}"
                            
                            # –ù–µ –¥–æ–±–∞–≤–ª—è–µ–º –æ—à–∏–±–∫–∏ –¥–ª—è –ø—Ä–æ–ø—É—â–µ–Ω–Ω—ã—Ö —Å—Ç—Ä–æ–∫ (–±–µ–∑ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞ –∏–ª–∏ –≤—Ä–µ–º–µ–Ω–∏)
                            if "–ù–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞ –Ω–µ —É–∫–∞–∑–∞–Ω" in str(e) or "–ù–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã–π –Ω–æ–º–µ—Ä –≤—ã–ª–µ—Ç–∞" in str(e):
                                # –≠—Ç–æ –Ω–µ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞, –ø—Ä–æ—Å—Ç–æ –ø—Ä–æ–ø—É—Å–∫–∞–µ–º —Å—Ç—Ä–æ–∫—É
                                skipped_errors += 1
                                if skipped_errors <= 20:
                                    logger.warning(f"–°—Ç—Ä–æ–∫–∞ {row_idx} –ø—Ä–æ–ø—É—â–µ–Ω–∞: {str(e)}")
                                continue
                            
                            # –î–æ–±–∞–≤–ª—è–µ–º —Ç–æ–ª—å–∫–æ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–∏–µ –æ—à–∏–±–∫–∏
                            skipped_errors += 1
                            if len(error_messages) < 200:
                                error_messages.append(error_msg)
                            if skipped_errors <= 20:
                                logger.warning(error_msg, exc_info=True)
                            
                            # –ü—Ä–æ–¥–æ–ª–∂–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É –¥–∞–∂–µ –ø—Ä–∏ –æ—à–∏–±–∫–∞—Ö
                        
                        # –ü—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ—Å–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å—Ç—Ä–æ–∫–∏, –Ω–µ –±—ã–ª –ª–∏ –ø—Ä–µ—Ä–≤–∞–Ω —Ü–∏–∫–ª –∏–∑-–∑–∞ –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫
                        if empty_rows_count >= MAX_EMPTY_ROWS:
                            break
                    
                    # –°–æ–∑–¥–∞–µ–º –æ—Å—Ç–∞–≤—à–∏–µ—Å—è –∑–∞–ø–∏—Å–∏ –≤ —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∞—Ö
                    # –ü–∏–ª–æ—Ç—ã —É–∂–µ —Å–æ–∑–¥–∞—é—Ç—Å—è —Å—Ä–∞–∑—É –ø—Ä–∏ –ø–µ—Ä–≤–æ–º –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–∏ —á–µ—Ä–µ–∑ get_or_create, –ø–æ—ç—Ç–æ–º—É –∏—Ö –Ω–µ –Ω—É–∂–Ω–æ —Å–æ–∑–¥–∞–≤–∞—Ç—å –∑–¥–µ—Å—å
                    
                    if new_target_types:
                        TargetType.objects.bulk_create(new_target_types.values(), ignore_conflicts=True)
                        # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                        target_types_cache = {tt.name.lower(): tt for tt in TargetType.objects.all()}
                    
                    if new_drones:
                        Drone.objects.bulk_create(new_drones.values(), ignore_conflicts=True)
                        # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                        # –°–æ–∑–¥–∞–µ–º –∫—ç—à –¥—Ä–æ–Ω–æ–≤ —Å –∫–ª—é—á–æ–º –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤ –¥–ª—è –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∏ X-51 –∏ X51
                        def get_drone_comparison_key(drone_name):
                            """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –¥—Ä–æ–Ω–æ–≤ (–±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤)"""
                            return re.sub(r'[-]', '', str(drone_name).lower().strip())
                        drones_cache = {get_drone_comparison_key(drone.name): drone for drone in Drone.objects.all()}
                        new_drones.clear()
                    
                    if new_explosive_types:
                        ExplosiveType.objects.bulk_create(new_explosive_types.values(), ignore_conflicts=True)
                        for et in new_explosive_types.values():
                            explosive_types_cache[et.name.lower()] = et
                    
                    if new_explosive_devices:
                        ExplosiveDevice.objects.bulk_create(new_explosive_devices.values(), ignore_conflicts=True)
                        for ed in new_explosive_devices.values():
                            explosive_devices_cache[ed.name.lower()] = ed
                    
                    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –æ—Å—Ç–∞–≤—à–∏–µ—Å—è –∑–∞–ø–∏—Å–∏ –ø–æ–ª–µ—Ç–æ–≤ (–ø–æ—Å–ª–µ–¥–Ω–∏–π –±–∞—Ç—á)
                    from django.db import transaction
                    if flights_to_create:
                        try:
                            with transaction.atomic():
                                # –°–Ω–∞—á–∞–ª–∞ —Å–æ–∑–¥–∞–µ–º —Å–ø—Ä–∞–≤–æ—á–Ω–∏–∫–∏, –µ—Å–ª–∏ –Ω—É–∂–Ω–æ
                                if new_target_types:
                                    TargetType.objects.bulk_create(new_target_types.values(), ignore_conflicts=True)
                                    # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                                    target_types_cache = {tt.name.lower(): tt for tt in TargetType.objects.all()}
                                
                                if new_drones:
                                    Drone.objects.bulk_create(new_drones.values(), ignore_conflicts=True)
                                    # –ü–µ—Ä–µ–∑–∞–≥—Ä—É–∂–∞–µ–º –∫—ç—à –∏–∑ –±–∞–∑—ã –¥–ª—è –ø–æ–ª—É—á–µ–Ω–∏—è –∞–∫—Ç—É–∞–ª—å–Ω—ã—Ö –Ω–æ—Ä–º–∞–ª–∏–∑–æ–≤–∞–Ω–Ω—ã—Ö –∑–Ω–∞—á–µ–Ω–∏–π
                                    # –°–æ–∑–¥–∞–µ–º –∫—ç—à –¥—Ä–æ–Ω–æ–≤ —Å –∫–ª—é—á–æ–º –±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤ –¥–ª—è –≥—Ä—É–ø–ø–∏—Ä–æ–≤–∫–∏ X-51 –∏ X51
                                    def get_drone_comparison_key(drone_name):
                                        """–í–æ–∑–≤—Ä–∞—â–∞–µ—Ç –∫–ª—é—á –¥–ª—è —Å—Ä–∞–≤–Ω–µ–Ω–∏—è –¥—Ä–æ–Ω–æ–≤ (–±–µ–∑ –¥–µ—Ñ–∏—Å–æ–≤)"""
                                        return re.sub(r'[-]', '', str(drone_name).lower().strip())
                                    drones_cache = {get_drone_comparison_key(drone.name): drone for drone in Drone.objects.all()}
                                    new_drones.clear()
                                
                                if new_explosive_types:
                                    ExplosiveType.objects.bulk_create(new_explosive_types.values(), ignore_conflicts=True)
                                    for et in new_explosive_types.values():
                                        explosive_types_cache[et.name.lower()] = et
                                
                                if new_explosive_devices:
                                    ExplosiveDevice.objects.bulk_create(new_explosive_devices.values(), ignore_conflicts=True)
                                    for ed in new_explosive_devices.values():
                                        explosive_devices_cache[ed.name.lower()] = ed
                                
                                if flights_to_create:
                                    logger.info(f"–°–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ –ø–æ—Å–ª–µ–¥–Ω–µ–≥–æ –±–∞—Ç—á–∞: {len(flights_to_create)} –∑–∞–ø–∏—Å–µ–π –¥–ª—è —Å–æ–∑–¥–∞–Ω–∏—è")
                                    # –†–∞–∑–±–∏–≤–∞–µ–º –Ω–∞ –º–µ–Ω—å—à–∏–µ –±–∞—Ç—á–∏ –¥–ª—è –∏–∑–±–µ–∂–∞–Ω–∏—è –Ω–µ—Ö–≤–∞—Ç–∫–∏ –ø–∞–º—è—Ç–∏
                                    batch_size = 50
                                    total_created_in_final = 0
                                    total_failed_in_final = 0
                                    for i in range(0, len(flights_to_create), batch_size):
                                        batch = flights_to_create[i:i + batch_size]
                                        try:
                                            created_objects = Flight.objects.bulk_create(batch, ignore_conflicts=True)
                                            # bulk_create —Å ignore_conflicts=True –≤–æ–∑–≤—Ä–∞—â–∞–µ—Ç –¢–û–õ–¨–ö–û —Å–æ–∑–¥–∞–Ω–Ω—ã–µ –æ–±—ä–µ–∫—Ç—ã
                                            created_count_in_sub_batch = len(created_objects) if created_objects else 0
                                            total_created_in_final += created_count_in_sub_batch
                                            
                                            # –ï—Å–ª–∏ —Å–æ–∑–¥–∞–Ω–æ –º–µ–Ω—å—à–µ, —á–µ–º –≤ –±–∞—Ç—á–µ - –∑–Ω–∞—á–∏—Ç –±—ã–ª–∏ –∫–æ–Ω—Ñ–ª–∏–∫—Ç—ã
                                            if created_count_in_sub_batch < len(batch):
                                                failed_count = len(batch) - created_count_in_sub_batch
                                                total_failed_in_final += failed_count
                                                # –ü—Ä–æ–±—É–µ–º —Å–æ–∑–¥–∞—Ç—å –ø–æ –æ–¥–Ω–æ–π –∑–∞–ø–∏—Å–∏ –¥–ª—è —Ç–µ—Ö, —á—Ç–æ –Ω–µ —Å–æ–∑–¥–∞–ª–∏—Å—å
                                                for flight in batch:
                                                    # –ü—Ä–æ–≤–µ—Ä—è–µ–º, –±—ã–ª –ª–∏ —Å–æ–∑–¥–∞–Ω —ç—Ç–æ—Ç –ø–æ–ª–µ—Ç (–ø—Ä–æ–≤–µ—Ä—è–µ–º –ø–æ –≤—Å–µ–º –∫–ª—é—á–µ–≤—ã–º –ø–æ–ª—è–º)
                                                    existing = Flight.objects.filter(
                                                        number=flight.number,
                                                        pilot=flight.pilot,
                                                        flight_date=flight.flight_date,
                                                        flight_time=flight.flight_time
                                                    ).first()
                                                    
                                                    if not existing:
                                                        try:
                                                            flight.save()
                                                            total_created_in_final += 1
                                                            total_failed_in_final -= 1
                                                            if total_failed_in_final <= 5:
                                                                logger.info(f"–£—Å–ø–µ—à–Ω–æ —Å–æ–∑–¥–∞–Ω –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname} –ø—Ä–∏ –ø–æ–≤—Ç–æ—Ä–Ω–æ–π –ø–æ–ø—ã—Ç–∫–µ (—Ñ–∏–Ω–∞–ª—å–Ω—ã–π –±–∞—Ç—á)")
                                                        except Exception as single_error:
                                                            if total_failed_in_final <= 10:
                                                                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname}: {single_error}")
                                                    else:
                                                        # –ó–∞–ø–∏—Å—å —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç - —ç—Ç–æ –Ω–æ—Ä–º–∞–ª—å–Ω–æ, –ø—Ä–æ—Å—Ç–æ –Ω–µ —Å—á–∏—Ç–∞–µ–º –∫–∞–∫ –æ—à–∏–±–∫—É
                                                        total_failed_in_final -= 1
                                                        if total_failed_in_final <= 5:
                                                            logger.debug(f"–ü–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname} —É–∂–µ —Å—É—â–µ—Å—Ç–≤—É–µ—Ç –≤ –ë–î (—Ñ–∏–Ω–∞–ª—å–Ω—ã–π –±–∞—Ç—á)")
                                        except Exception as batch_create_error:
                                            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ bulk_create —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞ {i//batch_size + 1}: {batch_create_error}", exc_info=True)
                                            # –ü—Ä–æ–±—É–µ–º —Å–æ–∑–¥–∞—Ç—å –ø–æ –æ–¥–Ω–æ–π –∑–∞–ø–∏—Å–∏
                                            for flight in batch:
                                                try:
                                                    flight.save()
                                                    total_created_in_final += 1
                                                except Exception as single_error:
                                                    total_failed_in_final += 1
                                                    if total_failed_in_final <= 10:
                                                        logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ–∑–¥–∞—Ç—å –ø–æ–ª–µ—Ç {flight.number} –¥–ª—è –ø–∏–ª–æ—Ç–∞ {flight.pilot.callname}: {single_error}")
                                    
                                    created_count += total_created_in_final
                                    if total_created_in_final != len(flights_to_create):
                                        logger.warning(f"–°–æ–∑–¥–∞–Ω–æ {total_created_in_final} –∏–∑ {len(flights_to_create)} –∑–∞–ø–∏—Å–µ–π –≤ —Ñ–∏–Ω–∞–ª—å–Ω–æ–º –±–∞—Ç—á–µ. –ù–µ —Å–æ–∑–¥–∞–Ω–æ: {total_failed_in_final}")
                                        self.message_user(request,
                                                          _(f"‚ö†Ô∏è –í —Ñ–∏–Ω–∞–ª—å–Ω–æ–º –±–∞—Ç—á–µ —Å–æ–∑–¥–∞–Ω–æ {total_created_in_final} –∏–∑ {len(flights_to_create)} –∑–∞–ø–∏—Å–µ–π. –ù–µ —Å–æ–∑–¥–∞–Ω–æ: {total_failed_in_final}"),
                                                          level=messages.WARNING)
                                    # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –≤ —Ñ–æ–Ω–æ–≤–æ–º —Ä–µ–∂–∏–º–µ –¥–ª—è —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞
                                    import threading
                                    def convert_coordinates_background_final():
                                        """–§–æ–Ω–æ–≤–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞"""
                                        try:
                                            from flights.models import Flight
                                            # –ù–∞—Ö–æ–¥–∏–º –≤—Å–µ –ø–æ–ª–µ—Ç—ã —Å –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç–∞–º–∏, –∫–æ—Ç–æ—Ä—ã–µ –µ—â–µ –Ω–µ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω—ã
                                            flights_to_convert = Flight.objects.filter(
                                                coordinates__isnull=False
                                            ).exclude(
                                                coordinates=''
                                            ).filter(
                                                lat_wgs84__isnull=True
                                            )[:200]  # –û–±—Ä–∞–±–∞—Ç—ã–≤–∞–µ–º –ø–æ 200 –∑–∞ —Ä–∞–∑ –¥–ª—è —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞
                                            
                                            converted_count = 0
                                            for flight in flights_to_convert:
                                                try:
                                                    # –ü—Ä–µ–æ–±—Ä–∞–∑—É–µ–º –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –∏ —Å–æ—Ö—Ä–∞–Ω—è–µ–º –≤ –ë–î
                                                    flight.get_coordinates_info_cached()
                                                    converted_count += 1
                                                except Exception as coord_error:
                                                    logger.warning(f"–û—à–∏–±–∫–∞ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è –ø–æ–ª–µ—Ç–∞ {flight.id}: {coord_error}")
                                            
                                            if converted_count > 0:
                                                logger.info(f"–ü—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –≤ —Ñ–æ–Ω–æ–≤–æ–º —Ä–µ–∂–∏–º–µ (—Ñ–∏–Ω–∞–ª—å–Ω—ã–π –±–∞—Ç—á): {converted_count} –ø–æ–ª–µ—Ç–æ–≤")
                                        except Exception as bg_error:
                                            logger.error(f"–û—à–∏–±–∫–∞ –≤ —Ñ–æ–Ω–æ–≤–æ–º –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç (—Ñ–∏–Ω–∞–ª—å–Ω—ã–π –±–∞—Ç—á): {bg_error}", exc_info=True)
                                    
                                    # –ó–∞–ø—É—Å–∫–∞–µ–º –≤ —Ñ–æ–Ω–æ–≤–æ–º –ø–æ—Ç–æ–∫–µ
                                    thread = threading.Thread(target=convert_coordinates_background_final, daemon=True)
                                    thread.start()
                                    logger.debug(f"–ó–∞–ø—É—â–µ–Ω–∞ —Ñ–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –ø—Ä–µ–æ–±—Ä–∞–∑–æ–≤–∞–Ω–∏—è –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –¥–ª—è —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞")
                                    
                                    # –ë–∞–∑–∞ –æ—á–∏—â–µ–Ω–∞, –ø–æ—ç—Ç–æ–º—É –Ω–µ –Ω—É–∂–Ω–æ –æ–±–Ω–æ–≤–ª—è—Ç—å existing_flights_set
                                    
                                    # –Ø–≤–Ω–∞—è –æ—á–∏—Å—Ç–∫–∞ –ø–∞–º—è—Ç–∏ –ø–æ—Å–ª–µ —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞
                                    import gc
                                    gc.collect()
                                    
                                    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ –ø–æ—Å–ª–µ —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞
                                    if import_progress:
                                        try:
                                            import_progress.last_processed_row = data_start_row + processed_row_count - 1
                                            import_progress.total_created = created_count
                                            import_progress.save(update_fields=['last_processed_row', 'total_created', 'last_import_date'])
                                            logger.info(f"–ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω –ø–æ—Å–ª–µ —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞: —Å—Ç—Ä–æ–∫–∞ {import_progress.last_processed_row}, —Å–æ–∑–¥–∞–Ω–æ {created_count}")
                                        except Exception as progress_error:
                                            logger.warning(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∏–º–ø–æ—Ä—Ç–∞: {progress_error}")
                                
                                if flights_to_create:
                                    self.message_user(request,
                                                      _(f"üíæ –°–æ—Ö—Ä–∞–Ω–µ–Ω –ø–æ—Å–ª–µ–¥–Ω–∏–π –±–∞—Ç—á: {len(flights_to_create)} —Å–æ–∑–¥–∞–Ω–æ"),
                                                      level=messages.INFO)
                        except Exception as final_error:
                            error_msg = f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞: {str(final_error)}"
                            error_messages.append(error_msg)
                            logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –±–∞—Ç—á–∞: {final_error}", exc_info=True)
                            self.message_user(request,
                                              _(f"‚ö†Ô∏è –û—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏ –ø–æ—Å–ª–µ–¥–Ω–µ–≥–æ –±–∞—Ç—á–∞ ({len(flights_to_create)} —Å–æ–∑–¥–∞–Ω–æ): {str(final_error)}"),
                                              level=messages.ERROR)

                    if error_messages:
                        for msg in error_messages[:5]:  # –ü–æ–∫–∞–∂–µ–º —Ç–æ–ª—å–∫–æ –ø–µ—Ä–≤—ã–µ 5 –æ—à–∏–±–æ–∫ –∏–∑ —ç—Ç–æ–≥–æ —Ñ–∞–π–ª–∞
                            self.message_user(request, msg, level=messages.WARNING)
                        if len(error_messages) > 5:
                            self.message_user(request,
                                              f"... –∏ –µ—â—ë {len(error_messages) - 5} –æ—à–∏–±–æ–∫ –≤ —Ñ–∞–π–ª–µ '{xlsx_file.name}'.",
                                              level=messages.WARNING)

                    total_created += created_count
                    total_errors.extend(error_messages)

                    # –§–æ—Ä–º–∏—Ä—É–µ–º –∏—Ç–æ–≥–æ–≤–æ–µ —Å–æ–æ–±—â–µ–Ω–∏–µ —Å –¥–µ—Ç–∞–ª—å–Ω–æ–π —Å—Ç–∞—Ç–∏—Å—Ç–∏–∫–æ–π
                    total_processed = created_count
                    total_skipped = skipped_no_flight_number + skipped_errors + skipped_no_date + skipped_no_pilot
                    summary_message = f"–ò–º–ø–æ—Ä—Ç –∏–∑ —Ñ–∞–π–ª–∞ '{xlsx_file.name}' –∑–∞–≤–µ—Ä—à–µ–Ω:\n"
                    summary_message += f"  - –°–æ–∑–¥–∞–Ω–æ –∑–∞–ø–∏—Å–µ–π: {created_count}\n"
                    summary_message += f"  - –£—Å–ø–µ—à–Ω–æ –æ–±—Ä–∞–±–æ—Ç–∞–Ω–æ —Å—Ç—Ä–æ–∫: {processed_successfully}\n"
                    summary_message += f"  - –û–±—Ä–∞–±–æ—Ç–∞–Ω–æ —Å—Ç—Ä–æ–∫ –∏–∑ —Ñ–∞–π–ª–∞: {processed_row_count} –∏–∑ {total_rows_to_process} (–¥–æ —Å—Ç—Ä–æ–∫–∏ {data_start_row + processed_row_count - 1})\n"
                    if empty_rows_count >= MAX_EMPTY_ROWS:
                        summary_message += f"  - ‚ö†Ô∏è –ò–º–ø–æ—Ä—Ç –æ—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω –Ω–∞ —Å—Ç—Ä–æ–∫–µ {data_start_row + processed_row_count - 1} –∏–∑-–∑–∞ {empty_rows_count} –ø—É—Å—Ç—ã—Ö —Å—Ç—Ä–æ–∫ –ø–æ–¥—Ä—è–¥\n"
                    elif processed_row_count >= total_rows_to_process:
                        summary_message += f"  - ‚úì –û–±—Ä–∞–±–æ—Ç–∞–Ω –≤–µ—Å—å —Ñ–∞–π–ª –¥–æ —Å—Ç—Ä–æ–∫–∏ {data_end_row}\n"
                    if total_skipped > 0:
                        summary_message += f"  - –ü—Ä–æ–ø—É—â–µ–Ω–æ —Å—Ç—Ä–æ–∫ –±–µ–∑ –Ω–æ–º–µ—Ä–∞ –≤—ã–ª–µ—Ç–∞: {skipped_no_flight_number}\n"
                        if skipped_no_date > 0:
                            summary_message += f"  - –ü—Ä–æ–ø—É—â–µ–Ω–æ —Å—Ç—Ä–æ–∫ –±–µ–∑ –¥–∞—Ç—ã: {skipped_no_date}\n"
                        if skipped_no_pilot > 0:
                            summary_message += f"  - –ü—Ä–æ–ø—É—â–µ–Ω–æ —Å—Ç—Ä–æ–∫ –±–µ–∑ –ø–∏–ª–æ—Ç–∞: {skipped_no_pilot}\n"
                        if skipped_errors > 0:
                            summary_message += f"  - –ü—Ä–æ–ø—É—â–µ–Ω–æ —Å—Ç—Ä–æ–∫ –∏–∑-–∑–∞ –æ—à–∏–±–æ–∫: {skipped_errors}\n"
                        summary_message += f"  - –í—Å–µ–≥–æ –ø—Ä–æ–ø—É—â–µ–Ω–æ: {total_skipped}"
                    
                    self.message_user(
                        request,
                        _(summary_message),
                        level=messages.SUCCESS
                    )
                    
                    # –û–±–Ω–æ–≤–ª—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ - –ø–æ–º–µ—á–∞–µ–º –∫–∞–∫ –∑–∞–≤–µ—Ä—à–µ–Ω–Ω—ã–π, –µ—Å–ª–∏ —Ñ–∞–π–ª –æ–±—Ä–∞–±–æ—Ç–∞–Ω –ø–æ–ª–Ω–æ—Å—Ç—å—é
                    if import_progress:
                        try:
                            final_row = data_start_row + processed_row_count - 1
                            is_completed = (processed_row_count >= total_rows_to_process) or (empty_rows_count >= MAX_EMPTY_ROWS)
                            import_progress.last_processed_row = final_row
                            import_progress.total_created = created_count
                            import_progress.is_completed = is_completed
                            import_progress.save(update_fields=['last_processed_row', 'total_created', 'is_completed', 'last_import_date'])
                            if is_completed:
                                logger.info(f"–ò–º–ø–æ—Ä—Ç —Ñ–∞–π–ª–∞ '{xlsx_file.name}' –∑–∞–≤–µ—Ä—à–µ–Ω –ø–æ–ª–Ω–æ—Å—Ç—å—é (—Å—Ç—Ä–æ–∫–∞ {final_row}, —Å–æ–∑–¥–∞–Ω–æ {created_count})")
                                
                                # –ê—Å–∏–Ω—Ö—Ä–æ–Ω–Ω–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –ø–æ—Å–ª–µ –∑–∞–≤–µ—Ä—à–µ–Ω–∏—è –∏–º–ø–æ—Ä—Ç–∞ (–≤ —Ñ–æ–Ω–æ–≤–æ–º –ø–æ—Ç–æ–∫–µ)
                                import threading
                                def process_coordinates_background():
                                    """–§–æ–Ω–æ–≤–∞—è —Ñ—É–Ω–∫—Ü–∏—è –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞"""
                                    try:
                                        logger.info("–ù–∞—á–∏–Ω–∞–µ—Ç—Å—è —Ñ–æ–Ω–æ–≤–∞—è –æ–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞...")
                                        flights_without_coords = Flight.objects.filter(
                                            coordinates__isnull=False
                                        ).exclude(
                                            coordinates=''
                                        ).filter(
                                            lat_wgs84__isnull=True
                                        )
                                        
                                        total_to_convert = flights_without_coords.count()
                                        if total_to_convert > 0:
                                            logger.info(f"–ù–∞–π–¥–µ–Ω–æ {total_to_convert} –ø–æ–ª–µ—Ç–æ–≤ –¥–ª—è –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç")
                                            
                                            def progress_callback(processed, total):
                                                if processed % 1000 == 0 or processed == total:
                                                    logger.info(f"–û–±—Ä–∞–±–æ—Ç–∞–Ω–æ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç: {processed}/{total}")
                                            
                                            success_count, error_count = Flight.batch_process_coordinates(
                                                queryset=flights_without_coords,
                                                batch_size=500,
                                                update_callback=progress_callback
                                            )
                                            
                                            logger.info(f"–û–±—Ä–∞–±–æ—Ç–∫–∞ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –∑–∞–≤–µ—Ä—à–µ–Ω–∞: —É—Å–ø–µ—à–Ω–æ {success_count}, –æ—à–∏–±–æ–∫ {error_count}")
                                    except Exception as bg_error:
                                        logger.error(f"–û—à–∏–±–∫–∞ –≤ —Ñ–æ–Ω–æ–≤–æ–π –æ–±—Ä–∞–±–æ—Ç–∫–µ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞: {bg_error}", exc_info=True)
                                
                                # –ó–∞–ø—É—Å–∫–∞–µ–º –æ–±—Ä–∞–±–æ—Ç–∫—É –≤ —Ñ–æ–Ω–æ–≤–æ–º –ø–æ—Ç–æ–∫–µ
                                thread = threading.Thread(target=process_coordinates_background, daemon=True)
                                thread.start()
                                logger.info("–ó–∞–ø—É—â–µ–Ω–∞ —Ñ–æ–Ω–æ–≤–∞—è –∑–∞–¥–∞—á–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ –∫–æ–æ—Ä–¥–∏–Ω–∞—Ç –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞")
                                
                                self.message_user(request,
                                                  _(f"‚úì –ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω. –§–∞–π–ª –æ–±—Ä–∞–±–æ—Ç–∞–Ω –ø–æ–ª–Ω–æ—Å—Ç—å—é. –ö–æ–æ—Ä–¥–∏–Ω–∞—Ç—ã –æ–±—Ä–∞–±–∞—Ç—ã–≤–∞—é—Ç—Å—è –≤ —Ñ–æ–Ω–æ–≤–æ–º —Ä–µ–∂–∏–º–µ."),
                                                  level=messages.SUCCESS)
                            else:
                                logger.info(f"–ò–º–ø–æ—Ä—Ç —Ñ–∞–π–ª–∞ '{xlsx_file.name}' –ø—Ä–∏–æ—Å—Ç–∞–Ω–æ–≤–ª–µ–Ω (—Å—Ç—Ä–æ–∫–∞ {final_row}, —Å–æ–∑–¥–∞–Ω–æ {created_count}). –ú–æ–∂–Ω–æ –ø—Ä–æ–¥–æ–ª–∂–∏—Ç—å –ø—Ä–∏ —Å–ª–µ–¥—É—é—â–µ–π –∑–∞–≥—Ä—É–∑–∫–µ.")
                                self.message_user(request,
                                                  _(f"üìå –ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω. –ü—Ä–∏ —Å–ª–µ–¥—É—é—â–µ–π –∑–∞–≥—Ä—É–∑–∫–µ —Ñ–∞–π–ª–∞ –∏–º–ø–æ—Ä—Ç –ø—Ä–æ–¥–æ–ª–∂–∏—Ç—Å—è —Å —Å—Ç—Ä–æ–∫–∏ {final_row + 1}."),
                                                  level=messages.INFO)
                        except Exception as progress_error:
                            logger.warning(f"–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–∏–Ω–∞–ª—å–Ω–æ–≥–æ –ø—Ä–æ–≥—Ä–µ—Å—Å–∞ –∏–º–ø–æ—Ä—Ç–∞: {progress_error}")
                    
                    # –ó–∞–∫—Ä—ã–≤–∞–µ–º workbook –∏ –æ—á–∏—â–∞–µ–º –ø–∞–º—è—Ç—å –ø–æ—Å–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞
                    if 'wb' in locals():
                        wb.close()
                        del wb
                    if 'ws' in locals():
                        del ws
                    import gc
                    gc.collect()
                    logger.info(f"–ü–∞–º—è—Ç—å –æ—á–∏—â–µ–Ω–∞ –ø–æ—Å–ª–µ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Ñ–∞–π–ª–∞ '{xlsx_file.name}'")

                except Exception as e:
                    import traceback
                    error_traceback = traceback.format_exc()
                    logger.error(f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∏–º–ø–æ—Ä—Ç–µ –∏–∑ —Ñ–∞–π–ª–∞ '{xlsx_file.name}': {str(e)}\n{error_traceback}")
                    
                    # –°–æ—Ö—Ä–∞–Ω—è–µ–º –ø—Ä–æ–≥—Ä–µ—Å—Å –¥–∞–∂–µ –ø—Ä–∏ –∫—Ä–∏—Ç–∏—á–µ—Å–∫–æ–π –æ—à–∏–±–∫–µ
                    if import_progress and 'processed_row_count' in locals():
                        try:
                            final_row = data_start_row + processed_row_count - 1
                            import_progress.last_processed_row = final_row
                            import_progress.total_created = created_count if 'created_count' in locals() else 0
                            import_progress.is_completed = False
                            import_progress.save(update_fields=['last_processed_row', 'total_created', 'is_completed', 'last_import_date'])
                            logger.info(f"–ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω –ø–æ—Å–ª–µ –æ—à–∏–±–∫–∏: —Å—Ç—Ä–æ–∫–∞ {final_row}, —Å–æ–∑–¥–∞–Ω–æ {import_progress.total_created}")
                            self.message_user(request,
                                              _(f"‚ö†Ô∏è –ò–º–ø–æ—Ä—Ç –ø—Ä–µ—Ä–≤–∞–Ω –∏–∑-–∑–∞ –æ—à–∏–±–∫–∏. –ü—Ä–æ–≥—Ä–µ—Å—Å —Å–æ—Ö—Ä–∞–Ω–µ–Ω (—Å—Ç—Ä–æ–∫–∞ {final_row}, —Å–æ–∑–¥–∞–Ω–æ {import_progress.total_created}). –ó–∞–≥—Ä—É–∑–∏—Ç–µ —Ñ–∞–π–ª —Å–Ω–æ–≤–∞, —á—Ç–æ–±—ã –ø—Ä–æ–¥–æ–ª–∂–∏—Ç—å."),
                                              level=messages.WARNING)
                        except Exception as progress_error:
                            logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å –ø—Ä–æ–≥—Ä–µ—Å—Å –ø—Ä–∏ –æ—à–∏–±–∫–µ: {progress_error}")
                    
                    self.message_user(request,
                                      _(f"–ö—Ä–∏—Ç–∏—á–µ—Å–∫–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ –∏–º–ø–æ—Ä—Ç–µ –∏–∑ —Ñ–∞–π–ª–∞ '{xlsx_file.name}': {str(e)}"),
                                      level=messages.ERROR)
                    # –ó–∞–∫—Ä—ã–≤–∞–µ–º workbook –∏ –æ—á–∏—â–∞–µ–º –ø–∞–º—è—Ç—å –¥–∞–∂–µ –ø—Ä–∏ –æ—à–∏–±–∫–µ
                    if 'wb' in locals():
                        try:
                            wb.close()
                        except:
                            pass
                        del wb
                    if 'ws' in locals():
                        del ws
                    import gc
                    gc.collect()

            final_message = f"–ò–º–ø–æ—Ä—Ç –≤—Å–µ—Ö —Ñ–∞–π–ª–æ–≤ –∑–∞–≤–µ—Ä—à–µ–Ω. –í—Å–µ–≥–æ: {total_created} —Å–æ–∑–¥–∞–Ω–æ."
            if total_errors:
                final_message += f" –í—Å–µ–≥–æ –æ—à–∏–±–æ–∫: {len(total_errors)}."

            message_level = messages.SUCCESS if not total_errors else messages.WARNING
            self.message_user(request, _(final_message), level=message_level)
            
            # –û—á–∏—â–∞–µ–º –∫—ç—à –¥–ª—è API –∫–∞—Ä—Ç—ã –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞
            from django.core.cache import cache
            try:
                # –û—á–∏—â–∞–µ–º –≤—Å–µ –∫–ª—é—á–∏ –∫—ç—à–∞, –Ω–∞—á–∏–Ω–∞—é—â–∏–µ—Å—è —Å 'flights_total:'
                # –ò—Å–ø–æ–ª—å–∑—É–µ–º delete_pattern –µ—Å–ª–∏ –¥–æ—Å—Ç—É–ø–µ–Ω (django-redis)
                if hasattr(cache, 'delete_pattern'):
                    cache.delete_pattern('rubicon:flights_total:*')
                else:
                    # –ê–ª—å—Ç–µ—Ä–Ω–∞—Ç–∏–≤–Ω—ã–π —Å–ø–æ—Å–æ–± - –æ—á–∏—Å—Ç–∏—Ç—å –≤–µ—Å—å –∫—ç—à
                    cache.clear()
                logger.info("–ö—ç—à –¥–ª—è API –∫–∞—Ä—Ç—ã –æ—á–∏—â–µ–Ω –ø–æ—Å–ª–µ –∏–º–ø–æ—Ä—Ç–∞")
            except Exception as cache_error:
                logger.warning(f"–ù–µ —É–¥–∞–ª–æ—Å—å –æ—á–∏—Å—Ç–∏—Ç—å –∫—ç—à: {cache_error}")

            return HttpResponseRedirect("../")

        context = dict(
            self.admin_site.each_context(request),
            title=_("–ò–º–ø–æ—Ä—Ç –≤—ã–ª–µ—Ç–æ–≤ –∏–∑ XLSX"),
        )
        return render(request, "admin/import_xlsx.html", context)
    
    def clear_database_view(self, request):
        """–û—á–∏—Å—Ç–∫–∞ –≤—Å–µ—Ö –¥–∞–Ω–Ω—ã—Ö, –∑–∞–≥—Ä—É–∂–µ–Ω–Ω—ã—Ö –∏–∑ Excel"""
        if request.method == 'POST':
            from django.db import transaction
            try:
                with transaction.atomic():
                    # –ü–æ–¥—Å—á–∏—Ç—ã–≤–∞–µ–º –∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –∑–∞–ø–∏—Å–µ–π –ø–µ—Ä–µ–¥ —É–¥–∞–ª–µ–Ω–∏–µ–º
                    flights_count = Flight.objects.count()
                    pilots_count = Pilot.objects.count()
                    drones_count = Drone.objects.count()
                    explosive_types_count = ExplosiveType.objects.count()
                    explosive_devices_count = ExplosiveDevice.objects.count()
                    target_types_count = TargetType.objects.count()
                    corrective_types_count = CorrectiveType.objects.count()
                    direction_types_count = DirectionType.objects.count()
                    import_progress_count = ImportProgress.objects.count()
                    
                    # –£–¥–∞–ª—è–µ–º –≤—Å–µ –¥–∞–Ω–Ω—ã–µ
                    Flight.objects.all().delete()
                    Pilot.objects.all().delete()
                    Drone.objects.all().delete()
                    ExplosiveType.objects.all().delete()
                    ExplosiveDevice.objects.all().delete()
                    TargetType.objects.all().delete()
                    CorrectiveType.objects.all().delete()
                    DirectionType.objects.all().delete()
                    ImportProgress.objects.all().delete()
                    
                    self.message_user(
                        request,
                        _(f"‚úÖ –ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –æ—á–∏—â–µ–Ω–∞! –£–¥–∞–ª–µ–Ω–æ:\n"
                          f"‚Ä¢ –í—ã–ª–µ—Ç—ã: {flights_count}\n"
                          f"‚Ä¢ –ü–∏–ª–æ—Ç—ã: {pilots_count}\n"
                          f"‚Ä¢ –î—Ä–æ–Ω—ã: {drones_count}\n"
                          f"‚Ä¢ –í–∏–¥—ã –ë–ü: {explosive_types_count}\n"
                          f"‚Ä¢ –í–∏–¥—ã –≤–∑—Ä—ã–≤–∞—Ç–µ–ª—è: {explosive_devices_count}\n"
                          f"‚Ä¢ –¢–∏–ø—ã —Ü–µ–ª–µ–π: {target_types_count}\n"
                          f"‚Ä¢ –¢–∏–ø—ã –∫–æ—Ä—Ä–µ–∫—Ç–∏—Ä–æ–≤–æ–∫: {corrective_types_count}\n"
                          f"‚Ä¢ –¢–∏–ø—ã –Ω–∞–ø—Ä–∞–≤–ª–µ–Ω–∏–π: {direction_types_count}\n"
                          f"‚Ä¢ –ü—Ä–æ–≥—Ä–µ—Å—Å –∏–º–ø–æ—Ä—Ç–∞: {import_progress_count}"),
                        level=messages.SUCCESS
                    )
                    logger.info(f"–ë–∞–∑–∞ –¥–∞–Ω–Ω—ã—Ö –æ—á–∏—â–µ–Ω–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª–µ–º {request.user.username}")
            except Exception as e:
                logger.error(f"–û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—á–∏—Å—Ç–∫–µ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö: {e}", exc_info=True)
                self.message_user(
                    request,
                    _(f"‚ùå –û—à–∏–±–∫–∞ –ø—Ä–∏ –æ—á–∏—Å—Ç–∫–µ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö: {str(e)}"),
                    level=messages.ERROR
                )
            
            return HttpResponseRedirect("../")
        
        # GET –∑–∞–ø—Ä–æ—Å - –ø–æ–∫–∞–∑—ã–≤–∞–µ–º —Å—Ç—Ä–∞–Ω–∏—Ü—É –ø–æ–¥—Ç–≤–µ—Ä–∂–¥–µ–Ω–∏—è
        context = {
            **self.admin_site.each_context(request),
            'title': _('–û—á–∏—Å—Ç–∫–∞ –±–∞–∑—ã –¥–∞–Ω–Ω—ã—Ö'),
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
            'flights_count': Flight.objects.count(),
            'pilots_count': Pilot.objects.count(),
            'drones_count': Drone.objects.count(),
            'explosive_types_count': ExplosiveType.objects.count(),
            'explosive_devices_count': ExplosiveDevice.objects.count(),
        }
        return render(request, "admin/clear_database.html", context)

